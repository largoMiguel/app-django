"""Exportación Excel de registros de asistencia."""
from __future__ import annotations

import io

from openpyxl import Workbook
from openpyxl.styles import Font

from .services import label_for_tipo


def build_registros_workbook(registros) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Asistencia"
    headers = [
        "ID",
        "Fecha y hora",
        "Cédula",
        "Funcionario",
        "Tipo",
        "Equipo",
        "Ubicación equipo",
    ]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    for r in registros:
        ws.append(
            [
                r.id,
                r.fecha_hora.strftime("%Y-%m-%d %H:%M:%S") if r.fecha_hora else "",
                r.funcionario.cedula,
                r.funcionario.nombre_completo,
                label_for_tipo(r.tipo),
                r.equipo.nombre,
                r.equipo.ubicacion or "",
            ]
        )

    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 40)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_diario_workbook(rows: list[dict], *, punches_per_day: int = 2) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Asistencia diaria"
    headers = ["Fecha", "Cédula", "Funcionario", "Entrada"]
    if punches_per_day == 4:
        headers += ["Salida almuerzo", "Retorno almuerzo"]
    headers += ["Salida", "Equipo", "Estado", "Marcaciones"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    def _hora(slot):
        return (slot or {}).get("hora") or ""

    for row in rows:
        line = [
            row.get("fecha", ""),
            row.get("funcionario_cedula", ""),
            row.get("funcionario_nombre", ""),
            _hora(row.get("entrada")),
        ]
        if punches_per_day == 4:
            line += [_hora(row.get("salida_almuerzo")), _hora(row.get("retorno_almuerzo"))]
        line += [
            _hora(row.get("salida")),
            row.get("equipo_nombre", ""),
            row.get("estado_label", ""),
            f"{row.get('marcaciones', 0)}/{row.get('marcaciones_esperadas', punches_per_day)}",
        ]
        ws.append(line)

    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 40)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
