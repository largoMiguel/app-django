"""Exportación Excel — FUID, TRD, transferencias."""
from __future__ import annotations

from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from apps.entities.models import Entity

from .models import Expediente, FuidRegistro, SerieDocumental, Transferencia

HEADER_FILL = PatternFill(start_color="3eafd4", end_color="3eafd4", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


def _write_sheet(ws, headers: list[str], rows: list[list]) -> None:
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
    for r_idx, row in enumerate(rows, 2):
        for c_idx, val in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.border = THIN_BORDER


def build_fuid_excel(entity: Entity) -> tuple[BytesIO, str]:
    wb = Workbook()
    ws = wb.active
    ws.title = "FUID"
    headers = [
        "CÓDIGO",
        "SERIE",
        "SUBSERIE",
        "UNIDAD DOCUMENTAL",
        "FECHA INICIAL",
        "FECHA FINAL",
        "FÍSICO",
        "ELECTRÓNICO",
        "CAJA",
        "CARPETA",
        "TOMO",
        "FOLIOS",
        "UBICACIÓN",
        "NOTAS",
    ]
    rows = []
    for r in FuidRegistro.objects.filter(entity=entity).order_by("id"):
        rows.append(
            [
                r.codigo,
                r.serie_nombre,
                r.subserie_nombre,
                r.unidad_documental,
                r.fecha_inicial.isoformat() if r.fecha_inicial else "",
                r.fecha_final.isoformat() if r.fecha_final else "",
                "Sí" if r.soporte_fisico else "No",
                "Sí" if r.soporte_electronico else "No",
                r.caja,
                r.carpeta,
                r.tomo,
                r.folios,
                r.ubicacion,
                r.notas,
            ]
        )
    _write_sheet(ws, headers, rows)
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf, f"FUID_{entity.code}_{entity.id}.xlsx"


def build_trd_excel(entity: Entity) -> tuple[BytesIO, str]:
    wb = Workbook()
    ws = wb.active
    ws.title = "TRD"
    headers = [
        "CÓDIGO",
        "NOMBRE",
        "ES SUBSERIE",
        "SERIE PADRE",
        "RET. GESTIÓN (años)",
        "RET. CENTRAL (años)",
        "DISPOSICIÓN",
        "PROCEDIMIENTO",
    ]
    rows = []
    for s in SerieDocumental.objects.filter(entity=entity, is_active=True).select_related("parent").order_by("codigo"):
        rows.append(
            [
                s.codigo,
                s.nombre,
                "Sí" if s.es_subserie else "No",
                s.parent.codigo if s.parent else "",
                s.retencion_gestion_anios,
                s.retencion_central_anios,
                s.disposicion_final,
                s.procedimiento,
            ]
        )
    _write_sheet(ws, headers, rows)
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf, f"TRD_{entity.code}_{entity.id}.xlsx"


def build_transferencias_excel(entity: Entity) -> tuple[BytesIO, str]:
    wb = Workbook()
    ws = wb.active
    ws.title = "Transferencias"
    headers = ["ID", "TIPO", "ESTADO", "ACTA", "EXPEDIENTES", "EJECUTADA", "NOTAS"]
    rows = []
    for t in Transferencia.objects.filter(entity=entity).prefetch_related("expedientes").order_by("-created_at"):
        rows.append(
            [
                t.id,
                t.get_tipo_display(),
                t.get_estado_display(),
                t.acta,
                ", ".join(e.codigo for e in t.expedientes.all()),
                t.ejecutada_at.isoformat() if t.ejecutada_at else "",
                t.notas,
            ]
        )
    _write_sheet(ws, headers, rows)
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf, f"Transferencias_{entity.code}_{entity.id}.xlsx"
