"""Servicios: almacenamiento, eventos, radicados, importación."""
from __future__ import annotations

import hashlib
import logging
import uuid
from io import BytesIO

from django.db import transaction
from django.utils import timezone
from django.utils.text import get_valid_filename
from openpyxl import load_workbook
from rest_framework.exceptions import ValidationError

from apps.common.file_delivery import signed_gestion_documental_url
from apps.common.storages import gestion_documental_file_storage
from apps.entities.models import Entity

from .models import (
    DisposicionFinal,
    EstadoExpediente,
    EstadoTransferencia,
    EtapaExpediente,
    EventoGD,
    Expediente,
    InstrumentoArchivistico,
    SerieDocumental,
    TipoEventoGD,
    TipoTransferencia,
    Transferencia,
)
from .storage_paths import acta_b2_key, documento_b2_key, instrumento_b2_key
from .validators import validate_archivo

logger = logging.getLogger(__name__)


def log_evento(entity: Entity, tipo: str, actor, detalle: dict | None = None) -> EventoGD:
    return EventoGD.objects.create(
        entity=entity,
        tipo=tipo,
        actor=actor if getattr(actor, "is_authenticated", False) else None,
        detalle=detalle or {},
    )


def next_codigo_expediente(entity_id: int) -> str:
    prefix = f"EXP-{timezone.localtime().strftime('%Y%m%d')}-"
    last = (
        Expediente.objects.filter(entity_id=entity_id, codigo__startswith=prefix)
        .order_by("-codigo")
        .values_list("codigo", flat=True)
        .first()
    )
    seq = 1
    if last:
        try:
            seq = int(last.rsplit("-", 1)[-1]) + 1
        except ValueError:
            seq = 1
    return f"{prefix}{seq:04d}"


def upload_to_b2(key: str, content: bytes, content_type: str = "") -> None:
    storage = gestion_documental_file_storage()
    storage.save(key, BytesIO(content))
    logger.debug("GD upload %s (%s bytes)", key, len(content))


def upload_instrumento_archivo(instrumento: InstrumentoArchivistico, uploaded_file, actor) -> InstrumentoArchivistico:
    validate_archivo(uploaded_file.name, uploaded_file.size)
    content = uploaded_file.read()
    key = instrumento_b2_key(instrumento.entity_id, instrumento.id, uploaded_file.name)
    upload_to_b2(key, content, uploaded_file.content_type or "")
    instrumento.b2_key = key
    instrumento.nombre_archivo = get_valid_filename(uploaded_file.name)
    instrumento.content_type = uploaded_file.content_type or ""
    instrumento.size = len(content)
    instrumento.save(
        update_fields=["b2_key", "nombre_archivo", "content_type", "size", "updated_at"]
    )
    log_evento(
        instrumento.entity,
        TipoEventoGD.INSTRUMENTO,
        actor,
        {"instrumento_id": instrumento.id, "archivo": instrumento.nombre_archivo},
    )
    return instrumento


def upload_documento_expediente(expediente: Expediente, uploaded_file, actor, **meta) -> "DocumentoExpediente":
    from .models import DocumentoExpediente

    validate_archivo(uploaded_file.name, uploaded_file.size)
    content = uploaded_file.read()
    sha = hashlib.sha256(content).hexdigest()
    key = documento_b2_key(expediente.entity_id, expediente.id, uploaded_file.name)
    upload_to_b2(key, content, uploaded_file.content_type or "")

    last_version = (
        DocumentoExpediente.objects.filter(expediente=expediente, nombre=uploaded_file.name)
        .order_by("-version")
        .values_list("version", flat=True)
        .first()
    )
    version = (last_version or 0) + 1

    doc = DocumentoExpediente.objects.create(
        expediente=expediente,
        entity=expediente.entity,
        nombre=get_valid_filename(uploaded_file.name),
        tipo_documental=meta.get("tipo_documental", ""),
        b2_key=key,
        content_type=uploaded_file.content_type or "",
        size=len(content),
        sha256=sha,
        version=version,
        folio_inicio=meta.get("folio_inicio"),
        folio_fin=meta.get("folio_fin"),
        fecha_documento=meta.get("fecha_documento"),
        uploaded_by=actor,
    )
    log_evento(
        expediente.entity,
        TipoEventoGD.DOCUMENTO,
        actor,
        {"expediente_id": expediente.id, "documento_id": doc.id, "nombre": doc.nombre},
    )
    return doc


def cerrar_expediente(expediente: Expediente, actor) -> Expediente:
    if expediente.estado == EstadoExpediente.CERRADO:
        return expediente
    expediente.estado = EstadoExpediente.CERRADO
    expediente.save(update_fields=["estado", "updated_at"])
    log_evento(
        expediente.entity,
        TipoEventoGD.CAMBIO_ESTADO,
        actor,
        {"expediente_id": expediente.id, "estado": EstadoExpediente.CERRADO},
    )
    return expediente


def ejecutar_transferencia(transferencia: Transferencia, actor) -> Transferencia:
    if transferencia.estado == EstadoTransferencia.EJECUTADA:
        return transferencia
    nueva_etapa = {
        TipoTransferencia.PRIMARIA: EtapaExpediente.CENTRAL,
        TipoTransferencia.SECUNDARIA: EtapaExpediente.HISTORICO,
    }[transferencia.tipo]
    with transaction.atomic():
        for exp in transferencia.expedientes.filter(entity=transferencia.entity):
            exp.etapa = nueva_etapa
            exp.save(update_fields=["etapa", "updated_at"])
        transferencia.estado = EstadoTransferencia.EJECUTADA
        transferencia.ejecutada_at = timezone.now()
        transferencia.save(update_fields=["estado", "ejecutada_at"])
    log_evento(
        transferencia.entity,
        TipoEventoGD.TRANSFERENCIA,
        actor,
        {"transferencia_id": transferencia.id, "tipo": transferencia.tipo, "etapa": nueva_etapa},
    )
    return transferencia


def importar_series_excel(entity: Entity, uploaded_file, actor, *, instrumento_id: int | None = None) -> dict:
    wb = load_workbook(uploaded_file, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    created = 0
    updated = 0
    instrumento = None
    if instrumento_id:
        instrumento = InstrumentoArchivistico.objects.filter(pk=instrumento_id, entity=entity).first()

    for row in rows:
        if not row or not row[0]:
            continue
        codigo = str(row[0]).strip()
        nombre = str(row[1] or "").strip() or codigo
        es_subserie = str(row[2] or "").strip().lower() in {"1", "true", "si", "sí", "subserie"}
        parent_codigo = str(row[3] or "").strip() if row[3] else ""
        ret_g = int(row[4] or 0) if row[4] is not None else 0
        ret_c = int(row[5] or 0) if row[5] is not None else 0
        disp = str(row[6] or "CT").strip().upper()[:2]
        if disp not in {c.value for c in DisposicionFinal}:
            disp = DisposicionFinal.CT
        parent = None
        if parent_codigo:
            parent = SerieDocumental.objects.filter(entity=entity, codigo=parent_codigo, es_subserie=False).first()
        obj, was_created = SerieDocumental.objects.update_or_create(
            entity=entity,
            codigo=codigo,
            defaults={
                "nombre": nombre,
                "es_subserie": es_subserie,
                "parent": parent,
                "retencion_gestion_anios": ret_g,
                "retencion_central_anios": ret_c,
                "disposicion_final": disp,
                "instrumento": instrumento,
                "is_active": True,
            },
        )
        if was_created:
            created += 1
        else:
            updated += 1

    log_evento(entity, TipoEventoGD.INSTRUMENTO, actor, {"import_series": {"created": created, "updated": updated}})
    return {"created": created, "updated": updated, "total_rows": len(rows)}


def archivo_url(b2_key: str, filename: str | None = None) -> str | None:
    if not b2_key:
        return None
    from django.conf import settings

    if settings.USE_B2_STORAGE and settings.FILE_DELIVERY_SIGNING_KEY:
        return signed_gestion_documental_url(b2_key, filename=filename)
    return None
