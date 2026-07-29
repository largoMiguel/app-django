"""Exportación Excel — módulo SECOP."""
from __future__ import annotations

import io
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill


HEADER_FILL = PatternFill("solid", fgColor="3EAFD4")
HEADER_FONT = Font(color="FFFFFF", bold=True)


def _autosize(ws) -> None:
    for col in ws.columns:
        letter = col[0].column_letter
        max_len = max(len(str(c.value or "")) for c in col)
        ws.column_dimensions[letter].width = min(max_len + 2, 48)


def build_contracts_excel(records: list[dict[str, Any]], titulo: str = "Contratos SECOP") -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Contratos"
    headers = [
        "Fuente",
        "Tipo",
        "Referencia",
        "Estado",
        "Modalidad",
        "Tipo contrato",
        "Proveedor",
        "Documento",
        "Valor",
        "Pagado",
        "Pendiente",
        "Fecha firma",
        "Fecha fin",
        "Objeto",
    ]
    ws.append(headers)
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")

    for r in records:
        ws.append([
            r.get("fuente"),
            r.get("tipo_registro"),
            r.get("referencia"),
            r.get("estado"),
            r.get("modalidad"),
            r.get("tipo"),
            r.get("proveedor"),
            r.get("documento_proveedor"),
            r.get("valor"),
            r.get("valor_pagado"),
            r.get("valor_pendiente"),
            r.get("fecha_firma"),
            r.get("fecha_fin"),
            (r.get("objeto") or "")[:500],
        ])
    _autosize(ws)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_alerts_excel(alerts: list[dict[str, Any]]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Alertas"
    headers = ["Severidad", "Código", "Título", "Mensaje", "Fuente", "Cantidad", "Valor implicado"]
    ws.append(headers)
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
    for a in alerts:
        ws.append([
            a.get("severidad"),
            a.get("codigo"),
            a.get("titulo"),
            a.get("mensaje"),
            a.get("fuente"),
            a.get("cantidad"),
            a.get("valor_implicado"),
        ])
    _autosize(ws)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
