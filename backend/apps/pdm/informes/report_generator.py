"""
Generador de Informes Institucionales para Plan de Desarrollo Municipal (PDM)
Basado en estándares colombianos de gestión pública territorial

Este módulo genera informes de gestión institucional alineados con:
- Constitución Política de Colombia de 1991
- Normatividad en planeación territorial
- Metodología General Ajustada (MGA) del DNP
- Agenda 2030 y Objetivos de Desarrollo Sostenible

Estructura del informe:
1. Portada institucional (con equipo de gobierno)
2. Introducción (marco legal y objetivo del informe)
3. Avance por líneas estratégicas (pilares del plan)
4. Avance por sectores MGA (áreas temáticas)
5. Avance por ODS (alineación con Agenda 2030)
6. Descripción de cumplimiento de metas
7. Ejecución del plan de acción por vigencia

Formato: PDF, DOCX, Excel
Estilo: Lenguaje técnico-administrativo, formal, tercera persona
"""

from reportlab.lib.pagesizes import letter
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER, TA_JUSTIFY, TA_LEFT, TA_RIGHT
from reportlab.platypus import (
    SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle,
    PageBreak, Image as RLImage
)
from io import BytesIO
from datetime import datetime
from typing import List, Dict, Any
import os
import base64
from collections import defaultdict

# Configurar matplotlib para uso en servidor (sin display)
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
from matplotlib import patches
import numpy as np
plt.rcParams['font.family'] = 'DejaVu Sans'

from apps.pdm.models import ActividadEstado, PDMEjecucionPresupuestal
from apps.pdm.metrics import actividad_aggs_for_productos, avance_general_producto, resumen_anio
from apps.common.report_cover import build_cover_flowables
from apps.common.report_theme import (
    BG_WHITE,
    LINE_MID,
    MPL_BAR,
    MPL_BAR_SOFT,
    MPL_LINE,
    MPL_TEXT,
    ROW_ALT,
    TEXT_DARK,
    banner_style_cmds,
    table_style_cmds,
)


def _normalize_compare_text(text: str) -> str:
    import re

    return re.sub(r"\s+", " ", (text or "").strip()).lower()


def _evidencia_descripcion_distinta(actividad, evidencia) -> bool:
    ev_desc = (getattr(evidencia, "descripcion", None) or "").strip()
    if not ev_desc:
        return False
    act_desc = (getattr(actividad, "descripcion", None) or "").strip()
    return _normalize_compare_text(ev_desc) != _normalize_compare_text(act_desc)


class PDMReportGenerator:
    """Generador de informes PDF con estructura general con mejoras de rendimiento y contenido"""

    REPORT_TITLE = "INFORME DE INDICADORES PLAN DE DESARROLLO"

    def __init__(
        self,
        entity,
        productos: List,
        actividades: List,
        anio: int,
        filtros: dict | None = None,
        usar_ia: bool = False,
        incluir_evidencias: bool = True,
        ai_analysis: dict | None = None,
        analytics: dict | None = None,
    ):
        self.entity = entity
        self.productos = productos
        self.actividades = actividades
        self.anio = anio
        self.filtros = filtros or {}
        self.usar_ia = usar_ia
        self.incluir_evidencias = incluir_evidencias
        self.ai_analysis = ai_analysis or {}
        self.analytics = analytics or {}
        self.use_template = False
        self.buffer = BytesIO()
        self.doc = None
        self.styles = None
        self.story = []
        self.page_number = 0
        self._cache_graficas = {}

    # Marco usable en páginas internas (letter + márgenes 0.8" + encabezado FM-PDM-001)
    MAX_FLOWABLE_WIDTH = 7.0 * inch
    MAX_FLOWABLE_HEIGHT = 9.0 * inch
    MAX_CHART_HEIGHT = 9.0 * inch
    EVIDENCIA_IMG_INCH = 2.4 * inch
    FRAME_SAFETY_PT = 18
    GRUPO_BLOCK_GAP = 0.15 * inch
    COLOR_PRIMARY = TEXT_DARK
    COLOR_HEADER = TEXT_DARK
    COLOR_ROW_ALT = ROW_ALT
    TABLE_WIDTH = 7.0 * inch

    def _institutional_styles(self) -> dict:
        if hasattr(self, "_inst_styles_cache"):
            return self._inst_styles_cache
        self._inst_styles_cache = {
            "banner": ParagraphStyle(
                "InstBanner",
                parent=self.styles["Normal"],
                textColor=colors.white,
                fontName="Helvetica-Bold",
                fontSize=10,
                alignment=TA_CENTER,
            ),
            "col_header": ParagraphStyle(
                "InstColHeader",
                parent=self.styles["Normal"],
                textColor=colors.white,
                fontName="Helvetica-Bold",
                fontSize=8,
                alignment=TA_CENTER,
                leading=10,
            ),
            "cell_center": ParagraphStyle(
                "InstCellCenter",
                parent=self.styles["Normal"],
                fontSize=9,
                alignment=TA_CENTER,
                leading=11,
            ),
            "cell_left": ParagraphStyle(
                "InstCellLeft",
                parent=self.styles["Normal"],
                fontSize=9,
                alignment=TA_LEFT,
                leading=11,
            ),
            "cell_bold": ParagraphStyle(
                "InstCellBold",
                parent=self.styles["Normal"],
                fontSize=9,
                fontName="Helvetica-Bold",
                alignment=TA_CENTER,
                leading=11,
            ),
        }
        return self._inst_styles_cache

    def _append_banner_table(self, title: str) -> None:
        st = self._institutional_styles()
        table = Table(
            [[Paragraph(f"<b>{title}</b>", st["banner"])]],
            colWidths=[self.TABLE_WIDTH],
            splitByRow=True,
        )
        table.setStyle(TableStyle(banner_style_cmds()))
        self.story.append(table)

    def _append_data_table(self, rows: list, col_widths: list) -> None:
        st = self._institutional_styles()
        table = Table(rows, colWidths=col_widths, splitByRow=True)
        table.setStyle(
            TableStyle(
                table_style_cmds(n_rows=len(rows), left_cols=tuple(range(len(col_widths))))
            )
        )
        self.story.append(table)

    def _update_flowable_limits(self, top_margin: float, bottom_margin: float) -> None:
        usable_height = letter[1] - top_margin - bottom_margin - self.FRAME_SAFETY_PT
        self._max_flowable_height = max(3.5 * inch, min(self.MAX_CHART_HEIGHT, usable_height))
        self._max_flowable_width = self.MAX_FLOWABLE_WIDTH

    def _sync_flowable_limits_from_doc(self) -> None:
        """Usa el alto/ancho real del frame ReportLab (más fiable con plantilla)."""
        frame_h = float(getattr(self.doc, "height", letter[1]))
        frame_w = float(getattr(self.doc, "width", letter[0] - inch))
        self._max_flowable_height = max(3.5 * inch, frame_h - self.FRAME_SAFETY_PT)
        self._max_flowable_width = min(self.MAX_FLOWABLE_WIDTH, frame_w - 4)

    def _chart_image_height(self, item_count: int) -> float:
        cap = getattr(self, "_max_flowable_height", self.MAX_CHART_HEIGHT)
        raw = max(item_count * 0.5 * inch, 3.2 * inch)
        return min(raw, cap)

    def _rl_image(self, img_buffer, width, height, kind=None):
        max_w = getattr(self, "_max_flowable_width", self.MAX_FLOWABLE_WIDTH)
        max_h = getattr(self, "_max_flowable_height", self.MAX_FLOWABLE_HEIGHT)
        scale = min(1.0, max_w / width, max_h / height) * 0.98
        width *= scale
        height *= scale
        img_buffer.seek(0)
        kwargs = {}
        if kind:
            kwargs["kind"] = kind
        return RLImage(img_buffer, width=width, height=height, **kwargs)

    def _evidencia_image(self, img_data: bytes) -> RLImage:
        """Tamaño fijo de evidencia (imagen ya normalizada en service)."""
        size = self.EVIDENCIA_IMG_INCH
        return RLImage(BytesIO(img_data), width=size, height=size)

    def _meta_programada_producto(self, producto) -> float:
        if self.anio == 0:
            return sum(
                float(getattr(producto, f"programacion_{y}", 0) or 0)
                for y in (2024, 2025, 2026, 2027)
            )
        return float(getattr(producto, f"programacion_{self.anio}", 0) or 0)

    def _resolve_plan_name(self) -> str:
        name = getattr(self.entity, "plan_name", None)
        if name and str(name).strip() and str(name).lower() != "none":
            return str(name).strip()
        for prod in self.productos:
            plan = getattr(prod, "nombre_plan", None)
            if plan and str(plan).strip() and str(plan).lower() != "none":
                return str(plan).strip()
        return "Plan de Desarrollo Municipal"

    def _aggs_for_producto(self, producto) -> dict:
        if not hasattr(self, "_aggs_by_clave"):
            claves = [p.clave_producto for p in self.productos]
            self._aggs_by_clave = actividad_aggs_for_productos(self.entity.id, claves)
        return self._aggs_by_clave.get(producto.clave_producto, {})

    def _avance_fisico_meta_grupo(self, productos: list) -> float:
        meta_prog = sum(self._meta_programada_producto(p) for p in productos)
        if meta_prog <= 0:
            return 0.0
        meta_ejec = sum(self._meta_ejecutada_producto(p) for p in productos)
        return min(100.0, (meta_ejec / meta_prog) * 100)

    def _meta_ejecutada_producto(self, producto) -> float:
        aggs = self._aggs_for_producto(producto)
        if self.anio == 0:
            return sum(
                resumen_anio(producto, y, aggs)["meta_ejecutada"] for y in (2024, 2025, 2026, 2027)
            )
        return resumen_anio(producto, self.anio, aggs)["meta_ejecutada"]

    def _presupuesto_producto(self, producto) -> tuple[float, float]:
        qs = PDMEjecucionPresupuestal.objects.filter(
            entity_id=self.entity.id,
            codigo_producto=producto.codigo_producto,
        )
        if self.anio != 0:
            qs = qs.filter(anio=self.anio)
        pto = sum(float(e.pto_definitivo or 0) for e in qs)
        pagos = sum(float(e.pagos or 0) for e in qs)
        return pto, pagos

    def _append_bar_chart(
        self,
        labels: list[str],
        values: list[float],
        title: str,
        cache_key: str,
    ) -> None:
        if not labels:
            return
        if cache_key in self._cache_graficas:
            self.story.append(self._cache_graficas[cache_key])
            return
        try:
            fig, ax = plt.subplots(
                figsize=(7, self._chart_image_height(len(labels)) / inch),
            )
            fig.patch.set_facecolor("white")
            bar_colors = [
                MPL_BAR if v >= 50 else MPL_BAR_SOFT
                for v in values
            ]
            y_pos = np.arange(len(labels))
            bars = ax.barh(y_pos, values, color=bar_colors, height=0.6, alpha=0.9)
            for bar, val in zip(bars, values):
                width = bar.get_width()
                ax.text(
                    width + 2,
                    bar.get_y() + bar.get_height() / 2,
                    f"{val:.1f}%",
                    ha="left",
                    va="center",
                    fontsize=10,
                    fontweight="bold",
                    color=MPL_TEXT,
                )
            ax.set_yticks(y_pos)
            ax.set_yticklabels(labels, fontsize=9, color=MPL_TEXT)
            ax.set_xlabel("Porcentaje de Avance (%)", fontsize=11, fontweight="bold", color=MPL_TEXT)
            ax.set_title(title, fontsize=13, fontweight="bold", color=MPL_TEXT, pad=20)
            ax.set_xlim(0, 110)
            ax.spines["top"].set_visible(False)
            ax.spines["right"].set_visible(False)
            ax.spines["left"].set_color(MPL_LINE)
            ax.spines["bottom"].set_color(MPL_LINE)
            ax.grid(axis="x", alpha=0.3, linestyle="--", color=MPL_LINE)
            ax.tick_params(colors=MPL_TEXT)
            ax.set_axisbelow(True)
            plt.tight_layout()
            img_buffer = BytesIO()
            plt.savefig(img_buffer, format="png", dpi=150, bbox_inches="tight", facecolor="white")
            img_buffer.seek(0)
            plt.close(fig)
            img = self._rl_image(img_buffer, 7 * inch, self._chart_image_height(len(labels)))
            self._cache_graficas[cache_key] = img
            self.story.append(img)
        except Exception as exc:
            print(f"   ❌ Error generando gráfica '{title}': {exc}")
        finally:
            plt.close("all")

    def _chart_values_for_group(self, productos: list, metric: str) -> float:
        if not productos:
            return 0.0
        if metric == "cumplimiento":
            return sum(self.calcular_avance_producto(p) for p in productos) / len(productos)
        if metric == "fisico_meta":
            return self._avance_fisico_meta_grupo(productos)
        if metric == "financiero":
            return sum(self.calcular_avance_financiero(p) for p in productos) / len(productos)
        return 0.0

    def _generate_tres_graficas_grupo(
        self,
        group_key,
        group_label: str,
        cache_prefix: str,
        label_max_len: int = 40,
    ) -> None:
        groups: dict[str, list] = defaultdict(list)
        for prod in self.productos:
            groups[group_key(prod)].append(prod)
        chart_specs = [
            ("cumplimiento", f"Avance de cumplimiento por {group_label}"),
            ("fisico_meta", f"Avance físico de metas por {group_label}"),
            ("financiero", f"Avance financiero por {group_label}"),
        ]
        for metric, title in chart_specs:
            labels: list[str] = []
            values: list[float] = []
            for name, prods in sorted(groups.items()):
                labels.append((name or "Sin asignar")[:label_max_len])
                values.append(self._chart_values_for_group(prods, metric))
            self._append_bar_chart(
                labels,
                values,
                title,
                f"grafica_{cache_prefix}_{metric}_{self.anio}",
            )
            self.story.append(Spacer(1, 0.08 * inch))

    def generate_grafica_por_dependencia(self) -> None:
        por_secretaria = (self.analytics or {}).get("por_secretaria") or []
        if not por_secretaria:
            return
        title_style = ParagraphStyle(
            "ChartSectionTitle",
            parent=self.styles["Heading2"],
            fontSize=12,
            textColor=TEXT_DARK,
            spaceAfter=4,
            spaceBefore=8,
            fontName="Helvetica-Bold",
        )
        self.story.append(
            Paragraph("AVANCE DE CUMPLIMIENTO POR DEPENDENCIA / RESPONSABLE", title_style)
        )
        labels = [item["secretaria"][:35] for item in por_secretaria]
        values = [float(item.get("avance_pct") or 0) for item in por_secretaria]
        self._append_bar_chart(
            labels,
            values,
            "Avance de cumplimiento por dependencia / responsable",
            f"grafica_secretaria_{self.anio}",
        )
        self.story.append(Spacer(1, 0.12 * inch))

    def _aggregate_grupo_productos(self, productos: list) -> dict:
        count = len(productos)
        if not count:
            return {
                "count": 0,
                "meta_total": 0.0,
                "indicadores": 0,
                "avance_fisico": 0.0,
                "avance_financiero": 0.0,
            }
        meta_total = sum(self._meta_programada_producto(p) for p in productos)
        avance_fisico = sum(self.calcular_avance_producto(p) for p in productos) / count
        avance_financiero = sum(self.calcular_avance_financiero(p) for p in productos) / count
        indicadores = len(
            {
                (p.indicador_producto_mga or p.personalizacion_indicador or p.codigo_producto or "").strip()
                for p in productos
            }
        )
        return {
            "count": count,
            "meta_total": meta_total,
            "indicadores": indicadores,
            "avance_fisico": avance_fisico,
            "avance_financiero": avance_financiero,
        }

    def _append_grupo_resumen_table(
        self,
        *,
        grupo_label: str,
        grupo_nombre: str,
        productos: list,
    ) -> None:
        """Tabla resumen consolidada por sector u ODS (sin detalle de productos)."""
        agg = self._aggregate_grupo_productos(productos)
        white_header = ParagraphStyle(
            "WhiteHeader",
            parent=self.styles["Normal"],
            textColor=colors.white,
            fontName="Helvetica-Bold",
            fontSize=8,
        )
        center_cell = ParagraphStyle(
            "CenterCell",
            parent=self.styles["Normal"],
            alignment=TA_CENTER,
            fontSize=8,
            leading=10,
        )
        justify_cell = ParagraphStyle(
            "JustifyCell",
            parent=self.styles["Normal"],
            alignment=TA_JUSTIFY,
            fontSize=8,
            leading=10,
        )

        stats_text = (
            f"<b>Productos:</b> {agg['count']}<br/>"
            f"<b>Meta programada:</b> {agg['meta_total']:,.0f}<br/>"
            f"<b>Indicadores:</b> {agg['indicadores']}"
        )
        nombre = (grupo_nombre or "Sin asignar").strip()
        if len(nombre) > 120:
            nombre = nombre[:117] + "..."

        data = [
            [
                Paragraph(f"<b>{grupo_label}</b>", white_header),
                Paragraph("<b>INFORMACIÓN ESTADÍSTICA</b>", white_header),
                Paragraph("<b>PRODUCTO(S)</b>", white_header),
                Paragraph("<b>INDICADOR DE PRODUCTO</b>", white_header),
                Paragraph("<b>AVANCE DEL PRODUCTO</b>", white_header),
                Paragraph("<b>AVANCE FINANCIERO</b>", white_header),
            ],
            [
                Paragraph(nombre, justify_cell),
                Paragraph(stats_text, center_cell),
                Paragraph(f"<b>{agg['count']}</b>", center_cell),
                Paragraph(f"<b>{agg['indicadores']}</b>", center_cell),
                Paragraph(f"<b>{agg['avance_fisico']:.1f}%</b>", center_cell),
                Paragraph(f"<b>{agg['avance_financiero']:.1f}%</b>", center_cell),
            ],
        ]
        table = Table(
            data,
            colWidths=[1.35 * inch, 1.55 * inch, 0.85 * inch, 1.05 * inch, 1.05 * inch, 1.15 * inch],
            splitByRow=True,
        )
        table.setStyle(
            TableStyle(
                table_style_cmds(
                    n_rows=len(data),
                    numeric_cols=(2, 3, 4, 5),
                    left_cols=(0, 1),
                )
            )
        )
        self.story.append(table)
        self.story.append(Spacer(1, self.GRUPO_BLOCK_GAP))
        
    def get_justify_style(self, fontSize=8):
        """Helper para crear estilos justificados reutilizables"""
        return ParagraphStyle(
            'JustifyStyle',
            parent=self.styles['Normal'],
            alignment=TA_JUSTIFY,
            fontSize=fontSize,
            leading=fontSize + 2
        )
        
    def add_header_footer(self, canvas, doc):
        """Encabezado y pie de página estándar"""
        canvas.saveState()
        
        # ENCABEZADO
        canvas.setFont('Helvetica', 8)
        # Código de formulario estándar
        canvas.drawString(0.5*inch, 10.5*inch, "FM-PDM-001")
        canvas.drawString(0.5*inch, 10.3*inch, "Versión: 1.0")
        
        # Número de página y título
        canvas.drawRightString(8*inch, 10.5*inch, f"Página {doc.page}")
        canvas.drawRightString(8*inch, 10.3*inch, self.REPORT_TITLE)
        
        # Línea separadora
        canvas.setStrokeColor(LINE_MID)
        canvas.setFillColor(TEXT_DARK)
        canvas.line(0.5*inch, 10.2*inch, 8*inch, 10.2*inch)
        
        # PIE DE PÁGINA
        canvas.setFont('Helvetica', 7)
        footer_text = f"Plan de Desarrollo Municipal - {self.entity.name}"
        canvas.drawCentredString(4.25*inch, 0.5*inch, footer_text)
        
        canvas.restoreState()
    
    def generate_portada(self):
        """Genera la portada institucional con estructura PQRS en escala de grises."""
        normal_style = ParagraphStyle(
            "CoverNormal",
            parent=self.styles["Normal"],
            fontSize=10,
            alignment=TA_JUSTIFY,
            spaceAfter=8,
        )
        equipo_style = ParagraphStyle(
            "EquipoGobierno",
            parent=self.styles["Normal"],
            fontSize=10,
            alignment=TA_CENTER,
            textColor=TEXT_DARK,
            spaceAfter=3,
            leading=11,
        )
        filter_style = ParagraphStyle(
            "FilterInfo",
            parent=self.styles["Normal"],
            fontSize=10,
            alignment=TA_CENTER,
            textColor=TEXT_DARK,
            spaceAfter=4,
        )
        alcaldia_style = ParagraphStyle(
            "Alcaldia",
            parent=self.styles["Normal"],
            fontSize=14,
            alignment=TA_CENTER,
            textColor=TEXT_DARK,
            spaceAfter=6,
        )

        anio_texto = "Vigencia 2024-2027" if self.anio == 0 else f"Vigencia {self.anio}"
        plan_name = self._resolve_plan_name().upper()

        extra: list = []
        extra.append(Spacer(1, 0.1 * inch))
        extra.append(Paragraph("Alcaldía Municipal", alcaldia_style))

        if self.filtros:
            filter_info = []
            if self.filtros.get("secretarias"):
                secs = ", ".join(self.filtros["secretarias"])
                filter_info.append(f"Secretarías: {secs}")
            if self.filtros.get("fecha_inicio") or self.filtros.get("fecha_fin"):
                inicio = self.filtros.get("fecha_inicio", "N/A")
                fin = self.filtros.get("fecha_fin", "N/A")
                filter_info.append(f"Período: {inicio} a {fin}")
            if self.filtros.get("estados"):
                estados = ", ".join(self.filtros["estados"])
                filter_info.append(f"Estados: {estados}")
            for info in filter_info:
                extra.append(Paragraph(info, filter_style))

        extra.append(Spacer(1, 0.15 * inch))
        extra.append(Paragraph("<b>Equipo de Gobierno Municipal</b>", equipo_style))
        for cargo in [
            "Alcalde Municipal",
            "Gestor(a) Social",
            "Jefe de Planeación Municipal",
            "Secretario de Gobierno",
            "Comisaría de Familia",
            "Inspector de Policía",
        ]:
            extra.append(Paragraph(cargo, equipo_style))

        top_spacer = 0.2 if self.use_template else 0.5
        cover = build_cover_flowables(
            title_line="Informe de Seguimiento",
            subtitle_line=f"Informe de Gestión / Rendición de Cuentas<br/>{plan_name}",
            entity_name=self.entity.name,
            period_text=anio_texto,
            normal_style=normal_style,
            top_spacer=top_spacer,
            extra_flowables=extra,
        )
        self.story.extend(cover)
    
    def generate_introduccion(self):
        """Genera la introducción institucional del informe"""
        title_style = ParagraphStyle(
            'SectionTitle',
            parent=self.styles['Heading1'],
            fontSize=16,
            textColor=TEXT_DARK,
            spaceAfter=16,
            spaceBefore=4,
            fontName='Helvetica-Bold'
        )
        
        self.story.append(Paragraph("INTRODUCCIÓN", title_style))
        
        anio_texto = "el cuatrienio 2024-2027" if self.anio == 0 else f"la vigencia {self.anio}"
        plan_name = self._resolve_plan_name()
        
        intro_text = f"""
        Los planes de desarrollo de las entidades territoriales constituyen la carta de navegación y el principal 
        instrumento de planeación para el desarrollo integral del territorio. Se configuran como herramientas de 
        carácter político y técnico, construidas mediante procesos democráticos y pluralistas, en las cuales se 
        materializan las decisiones, acciones, medios y recursos que orientan la gestión pública hacia el logro de 
        los objetivos de desarrollo territorial.
        <br/><br/>
        El Plan de Desarrollo Municipal "{plan_name}" fue adoptado mediante Acuerdo Municipal, en cumplimiento de 
        lo establecido en la Constitución Política de Colombia de 1991 y la normatividad vigente en materia de 
        planeación territorial. Este instrumento define las estrategias, programas y proyectos que guían la acción 
        gubernamental del municipio de {self.entity.name}.
        <br/><br/>
        El presente informe de gestión institucional tiene como objetivo presentar un balance integral de los 
        resultados alcanzados durante {anio_texto}, evidenciando el estado de ejecución de las metas programadas, 
        la inversión de recursos administrativos y financieros, así como el avance en el cumplimiento de los 
        compromisos adquiridos con la comunidad.
        <br/><br/>
        Este documento describe los logros y avances obtenidos, identifica las metas pendientes de cumplimiento, 
        y formula recomendaciones estratégicas para el fortalecimiento de la gestión pública municipal. La información 
        contenida se encuentra organizada por líneas estratégicas, sectores de intervención y su alineación con los 
        Objetivos de Desarrollo Sostenible (ODS) de la Agenda 2030.
        <br/><br/>
        El informe se estructura como un instrumento de rendición de cuentas ante la comunidad y de transparencia en 
        la gestión integral del territorio, enmarcado en los principios de eficiencia, eficacia y efectividad de la 
        administración pública.
        """
        
        justify_style = ParagraphStyle(
            'Justify',
            parent=self.styles['BodyText'],
            alignment=TA_JUSTIFY,
            fontSize=10
        )
        
        self.story.append(Paragraph(intro_text, justify_style))
        
        # RESUMEN EJECUTIVO con KPIs principales
        self.generar_resumen_ejecutivo()

        # Gráfica por dependencia / responsable
        self.generate_grafica_por_dependencia()
        
        # RESUMEN CON IA (si está habilitado)
        if self.usar_ia:
            self.generar_resumen_ia()
        
    
    def generar_resumen_ejecutivo(self):
        """Genera resumen ejecutivo con indicadores clave al inicio del informe"""
        try:
            title_style = ParagraphStyle(
                'ExecutiveTitle',
                parent=self.styles['Heading1'],
                fontSize=14,
                textColor=TEXT_DARK,
                spaceAfter=2,
                fontName='Helvetica-Bold'
            )
            
            self.story.append(Paragraph("RESUMEN EJECUTIVO", title_style))
            
            # Calcular KPIs generales
            total_productos = len(self.productos)
            # Total de actividades según filtro de año
            if self.anio == 0:
                total_actividades = len(self.actividades)
            else:
                total_actividades = sum(1 for act in self.actividades if act.anio == self.anio)
            
            # Avance promedio (solo considerar productos con programación en el año actual)
            suma_avances = 0
            productos_con_meta = 0
            total_meta = 0
            total_ejecutado = 0
            
            for prod in self.productos:
                # Obtener meta programada según año
                if self.anio == 0:
                    # Sumar todas las metas del cuatrienio
                    meta_anio = (
                        (getattr(prod, 'programacion_2024', 0) or 0) +
                        (getattr(prod, 'programacion_2025', 0) or 0) +
                        (getattr(prod, 'programacion_2026', 0) or 0) +
                        (getattr(prod, 'programacion_2027', 0) or 0)
                    )
                else:
                    # Solo meta del año específico
                    meta_anio = getattr(prod, f'programacion_{self.anio}', 0) or 0
                
                if meta_anio > 0:
                    productos_con_meta += 1
                    total_meta += meta_anio
                    suma_avances += self.calcular_avance_producto(prod)
                
            avance_promedio = suma_avances / productos_con_meta if productos_con_meta > 0 else 0

            pres = self.analytics.get("presupuesto", {})
            if pres.get("pto_definitivo"):
                avance_financiero_promedio = round(
                    (pres.get("pagos", 0) / pres["pto_definitivo"]) * 100, 1
                )
            else:
                avance_financiero_promedio = 0.0
            
            # Total presupuesto según año seleccionado
            total_presupuesto = 0
            for prod in self.productos:
                if self.anio == 0:
                    # Presupuesto del cuatrienio completo (suma de todos los años)
                    total_presupuesto += float(prod.total_2024 or 0)
                    total_presupuesto += float(prod.total_2025 or 0)
                    total_presupuesto += float(prod.total_2026 or 0)
                    total_presupuesto += float(prod.total_2027 or 0)
                else:
                    # Presupuesto del año específico
                    if self.anio == 2024:
                        total_presupuesto += float(prod.total_2024 or 0)
                    elif self.anio == 2025:
                        total_presupuesto += float(prod.total_2025 or 0)
                    elif self.anio == 2026:
                        total_presupuesto += float(prod.total_2026 or 0)
                    elif self.anio == 2027:
                        total_presupuesto += float(prod.total_2027 or 0)
            
            # TABLA DE KPIs PRINCIPALES
            white_bold = ParagraphStyle('WhiteBold', parent=self.styles['Normal'], 
                                       textColor=colors.white, fontName='Helvetica-Bold', fontSize=9)
            center_style = ParagraphStyle('Center', parent=self.styles['Normal'], 
                                         alignment=TA_CENTER, fontSize=10)
            
            kpis_data = [
                [
                    Paragraph('Total Productos', white_bold),
                    Paragraph('Avance Físico Promedio', white_bold),
                    Paragraph('Avance Financiero Promedio', white_bold),
                    Paragraph('Presupuesto Acumulado', white_bold)
                ],
                [
                    Paragraph(f'<b>{total_productos}</b>', center_style),
                    Paragraph(f'<b>{avance_promedio:.1f}%</b>', center_style),
                    Paragraph(f'<b>{avance_financiero_promedio:.1f}%</b>', center_style),
                    Paragraph(f'<b>${total_presupuesto:,.0f}</b>', center_style)
                ]
            ]
            
            kpis_table = Table(kpis_data, colWidths=[1.75*inch, 1.75*inch, 1.75*inch, 1.75*inch], splitByRow=True)
            kpis_table.setStyle(TableStyle(
                table_style_cmds(n_rows=len(kpis_data), numeric_cols=(0, 1, 2, 3))
            ))
            
            self.story.append(kpis_table)
            
            print("✅ Resumen ejecutivo generado")
            
        except Exception as e:
            print(f"⚠️ Error generando resumen ejecutivo: {e}")
            import traceback
            traceback.print_exc()
    
    def generar_resumen_ia(self):
        """
        Genera un resumen narrativo con IA (OpenAI) sobre el estado del PDM
        Mejora implementada: análisis inteligente opcional
        """
        try:
            from openai import OpenAI

            title_style = ParagraphStyle(
                'AITitle',
                parent=self.styles['Heading1'],
                fontSize=14,
                textColor=TEXT_DARK,
                spaceAfter=2,
                fontName='Helvetica-Bold'
            )
            
            self.story.append(Paragraph("ANÁLISIS NARRATIVO CON INTELIGENCIA ARTIFICIAL", title_style))

            resumen_ia = self._build_ai_narrative()
            if not resumen_ia:
                total_productos = len(self.productos)
                total_actividades = len([a for a in self.actividades if self.anio == 0 or a.anio == self.anio])
                actividades_completadas = len(
                    [
                        a
                        for a in self.actividades
                        if (self.anio == 0 or a.anio == self.anio) and a.estado == "COMPLETADA"
                    ]
                )

                suma_avances = sum(self.calcular_avance_producto(p) for p in self.productos)
                avance_promedio = suma_avances / total_productos if total_productos > 0 else 0

                anio_texto = "todos los años del cuatrienio 2024-2027" if self.anio == 0 else f"el año {self.anio}"

                prompt = f"""Eres un analista experto en gestión pública territorial colombiana.

Genera un análisis narrativo profesional y técnico del siguiente Plan de Desarrollo Municipal:

DATOS DEL INFORME:
- Entidad: {self.entity.name}
- Período: {anio_texto}
- Total de productos: {total_productos}
- Total de actividades: {total_actividades}
- Actividades completadas: {actividades_completadas} ({actividades_completadas/total_actividades*100 if total_actividades > 0 else 0:.1f}%)
- Avance físico promedio: {avance_promedio:.1f}%

El análisis debe:
1. Evaluar el nivel de cumplimiento general (excelente, bueno, regular, bajo)
2. Identificar fortalezas principales
3. Señalar áreas de mejora o riesgos
4. Dar recomendaciones estratégicas

Límite: 250 palabras. Usa lenguaje formal y técnico apropiado para gestión pública."""

                from django.conf import settings

                api_key = getattr(settings, "PQRS_REPORTS_OPENAI_API_KEY", "") or getattr(settings, "OPENAI_API_KEY", "")
                client = OpenAI(api_key=api_key)
                model = (
                    getattr(settings, "PQRS_REPORTS_OPENAI_MODEL", "")
                    or getattr(settings, "OPENAI_MODEL", "gpt-4o-mini")
                )

                response = client.chat.completions.create(
                    model=model,
                    messages=[
                        {
                            "role": "system",
                            "content": "Eres un experto en análisis de gestión pública territorial en Colombia.",
                        },
                        {"role": "user", "content": prompt},
                    ],
                    max_tokens=500,
                    temperature=0.7,
                )

                resumen_ia = response.choices[0].message.content
            
            # Agregar resumen al informe
            ia_style = ParagraphStyle(
                'IAText',
                parent=self.styles['BodyText'],
                alignment=TA_JUSTIFY,
                fontSize=10,
                spaceAfter=2,
                leftIndent=12,
                rightIndent=12,
                backColor=ROW_ALT,
                borderPadding=10
            )
            
            self.story.append(Paragraph(f"<i>{resumen_ia}</i>", ia_style))
            
            print("✅ Resumen con IA generado exitosamente")
            
        except ImportError:
            print("⚠️ OpenAI no está instalado. Saltando resumen con IA.")
        except Exception as e:
            print(f"⚠️ Error generando resumen con IA: {e}")
            # No lanzar excepción, solo advertir
            import traceback
            traceback.print_exc()
    
    def calcular_avance_producto(self, producto):
        """Avance físico alineado con metrics.resumen_anio / avance_general_producto."""
        try:
            aggs = self._aggs_for_producto(producto)
            if self.anio == 0:
                return avance_general_producto(producto, aggs)
            return resumen_anio(producto, self.anio, aggs)["porcentaje_avance"]
        except Exception:
            return 0.0
    
    def calcular_avance_financiero(self, producto) -> float:
        """
        Calcula el avance financiero real basado en la ejecución presupuestal
        Formula: (Pagos / Presupuesto Definitivo) * 100
        
        Si no hay datos de ejecución, retorna el avance físico como estimación
        """
        try:
            qs = PDMEjecucionPresupuestal.objects.filter(
                entity_id=self.entity.id,
                codigo_producto=producto.codigo_producto,
            )
            if self.anio != 0:
                qs = qs.filter(anio=self.anio)
            ejecuciones = list(qs)
            
            if not ejecuciones:
                return self.calcular_avance_producto(producto)
            
            total_definitivo = 0
            total_pagos = 0
            
            for ejecucion in ejecuciones:
                total_definitivo += float(ejecucion.pto_definitivo or 0)
                total_pagos += float(ejecucion.pagos or 0)
            
            if total_definitivo == 0:
                return self.calcular_avance_producto(producto)
            
            avance_financiero = (total_pagos / total_definitivo) * 100
            return min(100, max(0, avance_financiero))
            
        except Exception:
            return self.calcular_avance_producto(producto)
    
    def generate_grafica_moderna_lineas(self):
        """Genera gráfica moderna de avance por líneas estratégicas con caché"""
        lineas_data = {}
        for prod in self.productos:
            linea = prod.linea_estrategica or 'Sin Línea'
            if linea not in lineas_data:
                lineas_data[linea] = {'total': 0, 'suma_avance': 0}
            lineas_data[linea]['total'] += 1
            lineas_data[linea]['suma_avance'] += self.calcular_avance_producto(prod)

        lineas = []
        avances = []
        for linea, data in lineas_data.items():
            if data['total'] > 0:
                lineas.append(linea[:40])
                avances.append(data['suma_avance'] / data['total'])

        self._append_bar_chart(
            lineas,
            avances,
            'Avance por Línea Estratégica',
            f'grafica_lineas_{self.anio}',
        )

    def generate_grafica_moderna_sectores(self):
        """Genera tres gráficas de avance por sectores MGA."""
        self._generate_tres_graficas_grupo(
            group_key=lambda p: p.sector_mga or "Sin Sector",
            group_label="Sector MGA",
            cache_prefix="sectores",
        )

    def generate_grafica_moderna_ods(self):
        """Genera tres gráficas de avance por ODS."""
        self._generate_tres_graficas_grupo(
            group_key=lambda p: p.ods or "Sin ODS",
            group_label="Objetivo de Desarrollo Sostenible",
            cache_prefix="ods",
            label_max_len=45,
        )
    
    def generate_seccion_lineas(self):
        """Genera sección de avance por líneas estratégicas"""
        title_style = ParagraphStyle(
            'SectionTitle',
            parent=self.styles['Heading1'],
            fontSize=14,
            textColor=TEXT_DARK,
            spaceAfter=2,
            fontName='Helvetica-Bold'
        )
        
        self.story.append(Paragraph(
            "AVANCE DE CUMPLIMIENTO DE METAS DEL PLAN DE DESARROLLO POR LÍNEAS ESTRATÉGICAS",
            title_style
        ))
        
        # Definición conceptual
        justify_style = ParagraphStyle(
            'Justify',
            parent=self.styles['BodyText'],
            alignment=TA_JUSTIFY,
            fontSize=10,
            spaceAfter=2
        )
        
        concepto_lineas = """
        Las líneas estratégicas constituyen los pilares, ejes o dimensiones fundamentales sobre los cuales 
        se estructura el Plan de Desarrollo Municipal. Estas líneas orientan la gestión pública y la asignación 
        de recursos hacia el logro de resultados específicos en áreas prioritarias del desarrollo territorial, 
        garantizando coherencia entre los objetivos de gobierno y las necesidades de la población.
        """
        
        self.story.append(Paragraph(concepto_lineas, justify_style))
        self.story.append(Spacer(1, 0.1 * inch))
        
        # Generar gráfica moderna
        self.generate_grafica_moderna_lineas()
    
    def generate_tabla_productos(self):
        """Genera tabla detallada de productos por línea estratégica"""
        # Agrupar productos por línea
        productos_por_linea = {}
        for prod in self.productos:
            linea = prod.linea_estrategica or 'Sin Línea Estratégica'
            if linea not in productos_por_linea:
                productos_por_linea[linea] = []
            productos_por_linea[linea].append(prod)
        
        title_style = ParagraphStyle(
            'SectionTitle',
            parent=self.styles['Heading1'],
            fontSize=14,
            textColor=TEXT_DARK,
            spaceAfter=2,
            fontName='Helvetica-Bold'
        )
        
        self.story.append(Paragraph(
            "DESCRIPCIÓN DE CUMPLIMIENTO DE METAS PLAN DE DESARROLLO POR LÍNEAS ESTRATÉGICAS",
            title_style
        ))
        
        for linea, productos in productos_por_linea.items():
            # Encabezado de línea con texto blanco y fondo verde institucional
            linea_style = ParagraphStyle(
                'LineaTitle',
                parent=self.styles['Heading2'],
                fontSize=11,
                textColor=colors.white,
                backColor=TEXT_DARK,
                alignment=TA_CENTER,
                fontName='Helvetica-Bold',
                leftIndent=6,
                rightIndent=6,
                spaceAfter=2,
                spaceBefore=6
            )
            
            # Tabla de encabezado de línea (cell merged)
            header_data = [[Paragraph("LÍNEA ESTRATÉGICA", linea_style)]]
            header_table = Table(header_data, colWidths=[7*inch], splitByRow=True)
            header_table.setStyle(TableStyle(banner_style_cmds() + [
                ('TOPPADDING', (0, 0), (-1, -1), 4),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
            ]))
            
            self.story.append(header_table)
            
            # Descripción de la línea estratégica
            desc_linea_style = ParagraphStyle(
                'DescLinea',
                parent=self.styles['Normal'],
                fontSize=10,
                alignment=TA_CENTER,
                fontName='Helvetica-Bold',
                spaceAfter=2,
                spaceBefore=6
            )
            desc_table = Table([[Paragraph(linea.upper(), desc_linea_style)]], colWidths=[7*inch], splitByRow=True)
            desc_table.setStyle(TableStyle([
                ('LINEBELOW', (0, 0), (-1, -1), 0.5, LINE_MID),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('TOPPADDING', (0, 0), (-1, -1), 6),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
            ]))
            
            self.story.append(desc_table)
            
            # Estilo para encabezados con texto blanco
            white_header = ParagraphStyle('WhiteHeader', parent=self.styles['Normal'], textColor=colors.white, fontName='Helvetica-Bold', fontSize=9)
            
            # Tabla de productos
            data = [[
                Paragraph('<b>PRODUCTO(S)</b>', white_header),
                Paragraph('<b>INDICADOR DE PRODUCTO</b>', white_header),
                Paragraph('<b>AVANCE DEL PRODUCTO</b>', white_header),
                Paragraph('<b>AVANCE FINANCIERO</b>', white_header)
            ]]
            
            for prod in productos:
                producto_text = prod.producto_mga or prod.codigo_producto
                indicador_text = prod.indicador_producto_mga or prod.personalizacion_indicador or 'N/A'
                
                # Calcular avance físico usando nuestra función
                avance_fisico_porcentaje = self.calcular_avance_producto(prod)
                avance_fisico = f"{avance_fisico_porcentaje:.1f}%"
                
                # Calcular avance financiero REAL desde ejecución presupuestal
                avance_financiero_porcentaje = self.calcular_avance_financiero(prod)
                avance_financiero = f"{avance_financiero_porcentaje:.1f}%"
                
                # Crear estilo justificado para textos largos
                justify_cell = ParagraphStyle(
                    'JustifyCell',
                    parent=self.styles['Normal'],
                    alignment=TA_JUSTIFY,
                    fontSize=8,
                    leading=10
                )
                
                data.append([
                    Paragraph(producto_text, justify_cell),  # Texto justificado completo
                    Paragraph(indicador_text, justify_cell),  # Texto justificado completo
                    Paragraph(avance_fisico, self.styles['Normal']),
                    Paragraph(avance_financiero, self.styles['Normal'])
                ])
            
            table = Table(data, colWidths=[2.5*inch, 2.5*inch, 1*inch, 1*inch], splitByRow=True)
            table.setStyle(TableStyle(
                table_style_cmds(
                    n_rows=len(data),
                    numeric_cols=(2, 3),
                    left_cols=(0, 1),
                ) + [
                    ('FONTSIZE', (0, 0), (-1, 0), 9),
                    ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
                    ('TOPPADDING', (0, 0), (-1, 0), 4),
                    ('FONTSIZE', (0, 1), (-1, -1), 8),
                    ('TOPPADDING', (0, 1), (-1, -1), 6),
                    ('BOTTOMPADDING', (0, 1), (-1, -1), 6),
                ]
            ))
            
            self.story.append(table)
    
    def generate_seccion_sectores(self):
        """Genera sección de avance por sectores MGA"""
        title_style = ParagraphStyle(
            'SectionTitle',
            parent=self.styles['Heading1'],
            fontSize=14,
            textColor=TEXT_DARK,
            spaceAfter=2,
            spaceBefore=4,
            fontName='Helvetica-Bold'
        )
        
        self.story.append(Paragraph(
            "AVANCE DE CUMPLIMIENTO DE METAS DEL PLAN DE DESARROLLO POR SECTORES",
            title_style
        ))
        
        desc_text = """
        Los sectores constituyen las áreas temáticas de acción gubernamental mediante las cuales se organiza 
        la gestión pública municipal. Cada sector agrupa objetivos, metas y programas de inversión específicos 
        orientados a atender las necesidades y prioridades de la población en campos determinados del desarrollo 
        territorial. Esta clasificación sectorial permite una gestión integral y articulada de las políticas 
        públicas, facilitando el seguimiento y evaluación de resultados por áreas de intervención.
        <br/><br/>
        La organización sectorial corresponde a la Metodología General Ajustada (MGA) establecida por el 
        Departamento Nacional de Planeación (DNP) para la formulación y evaluación de proyectos de inversión 
        pública en Colombia.
        """
        
        justify_style = ParagraphStyle(
            'Justify',
            parent=self.styles['BodyText'],
            alignment=TA_JUSTIFY,
            fontSize=10,
            spaceAfter=2
        )
        
        self.story.append(Paragraph(desc_text, justify_style))
        self.story.append(Spacer(1, 0.1 * inch))
        
        # Generar gráficas de sectores (cumplimiento, físico y financiero)
        self.generate_grafica_moderna_sectores()
        
    
    def generate_seccion_ods(self):
        """Genera sección de avance por Objetivos de Desarrollo Sostenible"""
        title_style = ParagraphStyle(
            'SectionTitle',
            parent=self.styles['Heading1'],
            fontSize=14,
            textColor=TEXT_DARK,
            spaceAfter=2,
            fontName='Helvetica-Bold'
        )
        
        self.story.append(Paragraph(
            "AVANCE DE CUMPLIMIENTO DE METAS DEL PLAN DE DESARROLLO POR OBJETIVOS DE DESARROLLO SOSTENIBLE (ODS)",
            title_style
        ))
        
        desc_text = f"""
        Los Objetivos de Desarrollo Sostenible (ODS) constituyen un conjunto de 17 objetivos globales establecidos 
        por la Asamblea General de las Naciones Unidas en 2015, como parte integral de la Agenda 2030 para el 
        Desarrollo Sostenible. Esta agenda representa un compromiso universal de los Estados miembros para erradicar 
        la pobreza, proteger el planeta y garantizar que todas las personas gocen de paz, prosperidad y bienestar.
        <br/><br/>
        La República de Colombia, en cumplimiento de sus compromisos internacionales, ha incorporado los ODS en sus 
        instrumentos de planeación nacional y territorial. El Plan de Desarrollo Municipal de {self.entity.name} se encuentra 
        alineado con estos objetivos globales, contribuyendo desde el ámbito local al cumplimiento de las metas 
        establecidas en la Agenda 2030.
        <br/><br/>
        La presente sección evidencia la articulación entre las líneas estratégicas y sectores del Plan de Desarrollo 
        Municipal con los Objetivos de Desarrollo Sostenible, demostrando el compromiso de la administración municipal 
        con el desarrollo sostenible del territorio y el bienestar de sus habitantes.
        """
        
        justify_style = ParagraphStyle(
            'Justify',
            parent=self.styles['BodyText'],
            alignment=TA_JUSTIFY,
            fontSize=10,
            spaceAfter=2
        )
        
        self.story.append(Paragraph(desc_text, justify_style))
        self.story.append(Spacer(1, 0.1 * inch))
        
        # Generar gráficas de ODS (cumplimiento, físico y financiero)
        self.generate_grafica_moderna_ods()
        
    
    def _build_ai_narrative(self) -> str:
        conclusiones = (self.ai_analysis or {}).get("conclusiones", "").strip()
        recomendaciones = (self.ai_analysis or {}).get("recomendaciones") or []
        parts: list[str] = []
        if conclusiones:
            parts.append(conclusiones)
        if recomendaciones:
            rec_text = " ".join(f"{idx}. {rec}" for idx, rec in enumerate(recomendaciones, 1))
            parts.append(f"Recomendaciones estratégicas: {rec_text}")
        return " ".join(parts).strip()

    def generate_tabla_productos_detallada(self):
        """Genera tabla institucional por producto (formato oficial Sora-Boyacá)"""
        title_style = ParagraphStyle(
            'SectionTitle',
            parent=self.styles['Heading1'],
            fontSize=14,
            textColor=TEXT_DARK,
            spaceAfter=2,
            fontName='Helvetica-Bold'
        )
        
        anio_vigencia = "EL CUATRIENIO 2024-2027" if self.anio == 0 else f"LA VIGENCIA {self.anio}"
        
        self.story.append(Paragraph(
            f"EJECUCIÓN DEL PLAN DE ACCIÓN - {anio_vigencia}",
            title_style
        ))
        
        # Agrupar actividades por producto
        actividades_por_producto = defaultdict(list)
        for act in self.actividades:
            if self.anio == 0 or act.anio == self.anio:
                actividades_por_producto[act.clave_producto].append(act)
        
        # Procesar cada producto (SIN LÍMITE - mejora implementada)
        st = self._institutional_styles()
        
        for prod in self.productos:
            actividades = actividades_por_producto.get(prod.clave_producto, [])
            producto_nombre = prod.producto_mga or prod.codigo_producto
            indicador_nombre = prod.indicador_producto_mga or prod.personalizacion_indicador or "N/A"
            avance_fisico = self.calcular_avance_producto(prod)
            avance_financiero = self.calcular_avance_financiero(prod)
            meta_ejecutada = self._meta_ejecutada_producto(prod)
            meta_programada = self._meta_programada_producto(prod)
            pto_definitivo, pagos = self._presupuesto_producto(prod)
            responsable = prod.responsable_secretaria.nombre if prod.responsable_secretaria else "N/A"
            anio_vigencia = "2024-2027" if self.anio == 0 else str(self.anio)
            meta_prog_label = (
                "META FÍSICA PROGRAMADA (CUATRIENIO)"
                if self.anio == 0
                else f"META FÍSICA PROGRAMADA ({self.anio})"
            )

            linea_nombre = (prod.linea_estrategica or "SIN LÍNEA").upper()
            linea_table = Table(
                [
                    [Paragraph("LÍNEA ESTRATÉGICA", st["banner"])],
                    [Paragraph(f"<b>{linea_nombre}</b>", st["cell_center"])],
                ],
                colWidths=[self.TABLE_WIDTH],
                splitByRow=True,
            )
            linea_table.setStyle(
                TableStyle(
                    banner_style_cmds()
                    + [
                        ("BACKGROUND", (0, 1), (-1, 1), ROW_ALT),
                        ("TEXTCOLOR", (0, 1), (-1, 1), TEXT_DARK),
                        ("LINEBELOW", (0, 0), (-1, -1), 0.5, LINE_MID),
                        ("TOPPADDING", (0, 0), (-1, -1), 5),
                        ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
                    ]
                )
            )
            self.story.append(linea_table)
            
            resumen_rows = [
                [
                    Paragraph("<b>INDICADOR DE PRODUCTO</b>", st["col_header"]),
                    Paragraph("<b>PRODUCTO(S)</b>", st["col_header"]),
                    Paragraph("", st["col_header"]),
                ],
                [
                    Paragraph(indicador_nombre, st["cell_left"]),
                    Paragraph(producto_nombre, st["cell_left"]),
                    Paragraph("", st["cell_left"]),
                ],
                [
                    Paragraph("<b>AVANCE FÍSICO</b>", st["col_header"]),
                    Paragraph("<b>AVANCE FINANCIERO</b>", st["col_header"]),
                    Paragraph("<b>META EJECUTADA</b>", st["col_header"]),
                ],
                [
                    Paragraph(f"{avance_fisico:.0f}%", st["cell_bold"]),
                    Paragraph(f"{avance_financiero:.0f}%", st["cell_bold"]),
                    Paragraph(f"{meta_ejecutada:g}", st["cell_bold"]),
                ],
                [
                    Paragraph(f"<b>{meta_prog_label}</b>", st["col_header"]),
                    Paragraph("<b>PRESUPUESTO TOTAL</b>", st["col_header"]),
                    Paragraph("<b>PAGOS EJECUTADOS</b>", st["col_header"]),
                ],
                [
                    Paragraph(f"{meta_programada:g}", st["cell_bold"]),
                    Paragraph(f"${pto_definitivo:,.0f}", st["cell_bold"]),
                    Paragraph(f"${pagos:,.0f}", st["cell_bold"]),
                ],
                [
                    Paragraph("<b>RESPONSABLE</b>", st["col_header"]),
                    Paragraph(f"<b>{responsable.upper()}</b>", st["cell_bold"]),
                    Paragraph("", st["cell_left"]),
                ],
            ]
            resumen_table = Table(
                resumen_rows,
                colWidths=[2.33 * inch, 2.34 * inch, 2.33 * inch],
                splitByRow=True,
            )
            resumen_table.setStyle(
                TableStyle(
                    [
                        ("BACKGROUND", (0, 0), (-1, 0), self.COLOR_HEADER),
                        ("BACKGROUND", (0, 2), (-1, 2), self.COLOR_HEADER),
                        ("BACKGROUND", (0, 4), (-1, 4), self.COLOR_HEADER),
                        ("BACKGROUND", (0, 6), (0, 6), self.COLOR_HEADER),
                        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                        ("TEXTCOLOR", (0, 2), (-1, 2), colors.white),
                        ("TEXTCOLOR", (0, 4), (-1, 4), colors.white),
                        ("TEXTCOLOR", (0, 6), (0, 6), colors.white),
                        ("SPAN", (1, 0), (2, 0)),
                        ("SPAN", (1, 1), (2, 1)),
                        ("SPAN", (1, 6), (2, 6)),
                        ("BACKGROUND", (0, 1), (-1, 1), self.COLOR_ROW_ALT),
                        ("BACKGROUND", (0, 3), (-1, 3), BG_WHITE),
                        ("BACKGROUND", (0, 5), (-1, 5), ROW_ALT),
                        ("BACKGROUND", (0, 6), (0, 6), self.COLOR_HEADER),
                        ("BACKGROUND", (1, 6), (2, 6), BG_WHITE),
                        ("TEXTCOLOR", (1, 6), (2, 6), TEXT_DARK),
                        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                        ("LINEBELOW", (0, 0), (-1, -1), 0.5, LINE_MID),
                        ("TOPPADDING", (0, 0), (-1, -1), 5),
                        ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
                        ("LEFTPADDING", (0, 0), (-1, -1), 6),
                        ("RIGHTPADDING", (0, 0), (-1, -1), 6),
                    ]
                )
            )
            self.story.append(resumen_table)

            self._append_banner_table(f"EJECUCIÓN PLAN DE ACCIÓN VIGENCIA {anio_vigencia}")

            actividades_rows = [[
                Paragraph("<b>Meta y/o Actividades</b>", st["col_header"]),
                Paragraph("<b>Descripción</b>", st["col_header"]),
                Paragraph("<b>Ejecutado</b>", st["col_header"]),
            ]]
            if actividades:
                for act in actividades:
                    nombre_actividad = act.nombre or "Sin nombre"
                    if len(nombre_actividad) > 400:
                        nombre_actividad = nombre_actividad[:397] + "..."
                    descripcion = (act.descripcion or "").strip()
                    if len(descripcion) > 4000:
                        descripcion = descripcion[:3997] + "..."
                    meta_act = float(act.meta_ejecutar or 0)
                    actividades_rows.append([
                        Paragraph(f"<b>{nombre_actividad}</b>", st["cell_left"]),
                        Paragraph(descripcion or "—", st["cell_left"]),
                        Paragraph(f"{meta_act:g}", st["cell_bold"]),
                    ])
            else:
                actividades_rows.append([
                    Paragraph("Sin actividades registradas", st["cell_left"]),
                    Paragraph("—", st["cell_left"]),
                    Paragraph("—", st["cell_center"]),
                ])

            actividades_table = Table(
                actividades_rows,
                colWidths=[2.5 * inch, 3.25 * inch, 1.25 * inch],
                splitByRow=True,
            )
            actividades_table.setStyle(
                TableStyle(
                    table_style_cmds(
                        n_rows=len(actividades_rows),
                        numeric_cols=(2,),
                        left_cols=(0, 1),
                    )
                    + [
                        ("VALIGN", (0, 0), (-1, -1), "TOP"),
                    ]
                )
            )
            self.story.append(actividades_table)

            # REGISTRO DE EVIDENCIA + IMÁGENES
            evidencias_encontradas = False
            
            # OPTIMIZACIÓN: Solo verificar si hay evidencias (sin cargar imágenes aún)
            actividades_con_evidencia = [act for act in actividades if hasattr(act, 'tiene_evidencia') and act.tiene_evidencia]
            
            if actividades_con_evidencia and self.incluir_evidencias:
                evidencias_dict = {}
                for act in actividades_con_evidencia:
                    evidencia = getattr(act, "evidencia", None)
                    if evidencia is not None:
                        evidencias_dict[act.id] = evidencia

                evidencias_a_mostrar = [
                    act
                    for act in actividades_con_evidencia
                    if act.id in evidencias_dict
                    and (
                        getattr(evidencias_dict[act.id], "imagenes", None)
                        or (getattr(evidencias_dict[act.id], "url_evidencia", None) or "").strip()
                    )
                ]

                if evidencias_a_mostrar:
                    evidencias_encontradas = True
                    self._append_banner_table("REGISTRO DE EVIDENCIAS")

                    for num_evidencia, actividad in enumerate(evidencias_a_mostrar, 1):
                        evidencia = evidencias_dict[actividad.id]

                        actividad_nombre = actividad.nombre[:400] if len(actividad.nombre) > 400 else actividad.nombre
                        self.story.append(Paragraph(
                            f"<b>Actividad {num_evidencia}:</b> {actividad_nombre}",
                            ParagraphStyle('EvidenciaTitle', parent=self.styles['Normal'], fontSize=9, textColor=TEXT_DARK)
                        ))

                        if _evidencia_descripcion_distinta(actividad, evidencia):
                            ev_desc = (evidencia.descripcion or "").strip()
                            if len(ev_desc) > 4000:
                                ev_desc = ev_desc[:3997] + "..."
                            self.story.append(Paragraph(
                                f"<i>Descripción evidencia:</i> {ev_desc}",
                                ParagraphStyle('EvidenciaDesc', parent=self.styles['Normal'], fontSize=8, textColor=TEXT_DARK)
                            ))

                        url_ev = (getattr(evidencia, "url_evidencia", None) or "").strip()
                        if url_ev:
                            safe_url = url_ev.replace("&", "&amp;")
                            self.story.append(Paragraph(
                                f'<link href="{safe_url}" color="blue">Ver evidencia externa: {safe_url}</link>',
                                ParagraphStyle('EvidenciaUrl', parent=self.styles['Normal'], fontSize=8, textColor=TEXT_DARK)
                            ))

                        imagenes_cargadas = []
                        if evidencia.imagenes and isinstance(evidencia.imagenes, list):
                            for idx, img_base64 in enumerate(evidencia.imagenes):
                                try:
                                    if img_base64.startswith('data:image'):
                                        img_base64 = img_base64.split(',')[1]

                                    img_data = base64.b64decode(img_base64)
                                    img = self._evidencia_image(img_data)
                                    imagenes_cargadas.append(img)
                                    print(f"      ✅ Evidencia {num_evidencia} - Imagen {idx+1} agregada")
                                except Exception as e:
                                    print(f"      ⚠️ Error evidencia {num_evidencia} imagen {idx+1}: {e}")

                        if imagenes_cargadas:
                            grid_data = []
                            for i in range(0, len(imagenes_cargadas), 2):
                                row = imagenes_cargadas[i:i+2]
                                if len(row) == 1:
                                    row.append('')
                                grid_data.append(row)

                            img_table = Table(
                                grid_data,
                                colWidths=[self.EVIDENCIA_IMG_INCH + 0.1 * inch, self.EVIDENCIA_IMG_INCH + 0.1 * inch],
                                splitByRow=True,
                            )
                            img_table.setStyle(TableStyle([
                                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                                ('TOPPADDING', (0, 0), (-1, -1), 5),
                                ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
                            ]))

                            self.story.append(img_table)

                        self.story.append(Spacer(1, 0.15 * inch))
            
            if not evidencias_encontradas:
                self._append_banner_table("REGISTRO DE EVIDENCIA")
            
            # Separador entre productos (después de evidencias y antes del siguiente bloque)
            if idx < total_productos:
                self.story.append(Spacer(1, 0.35 * inch))
    
    def _build_content_pdf(
        self,
        *,
        top_margin: float,
        bottom_margin: float,
        use_template: bool,
    ) -> bytes:
        """Genera el PDF de contenido con márgenes según plantilla o FM-PDM-001."""
        self.use_template = use_template
        self.buffer = BytesIO()
        self.story = []
        self._cache_graficas = {}
        self._update_flowable_limits(top_margin, bottom_margin)

        self.doc = SimpleDocTemplate(
            self.buffer,
            pagesize=letter,
            rightMargin=0.5 * inch,
            leftMargin=0.5 * inch,
            topMargin=top_margin,
            bottomMargin=bottom_margin,
        )
        self._sync_flowable_limits_from_doc()
        self.styles = getSampleStyleSheet()

        print("  ├─ Portada")
        self.generate_portada()
        print("  ├─ Introducción")
        self.generate_introduccion()
        print("  ├─ Líneas Estratégicas")
        self.generate_seccion_lineas()
        print("  ├─ Sectores MGA")
        self.generate_seccion_sectores()
        print("  ├─ Objetivos de Desarrollo Sostenible")
        self.generate_seccion_ods()
        print("  ├─ Detalle de Productos con Actividades y Evidencias")
        self.generate_tabla_productos_detallada()

        print("  └─ Construyendo PDF...")
        if use_template:
            self.doc.build(self.story)
        else:
            self.doc.build(
                self.story,
                onFirstPage=self.add_header_footer,
                onLaterPages=self.add_header_footer,
            )

        pdf_bytes = self.buffer.getvalue()
        self.buffer.close()
        return pdf_bytes

    def generate(self) -> bytes:
        """Genera el PDF completo y retorna los bytes"""
        try:
            print("📄 Generando informe PDM en PDF...")
            return self._build_content_pdf(
                top_margin=0.8 * inch,
                bottom_margin=0.8 * inch,
                use_template=False,
            )
            
        except Exception as e:
            print(f"❌ Error generando PDF: {e}")
            import traceback
            traceback.print_exc()
            raise

    def generate_pdf(self) -> BytesIO:
        """Genera PDF con membrete institucional de la entidad cuando existe plantilla."""
        from apps.common.pdf_template import (
            apply_template_overlay,
            detect_template_margins,
            load_entity_template,
        )

        template_pdf_bytes = load_entity_template(self.entity)
        top_margin = 0.8 * inch
        bottom_margin = 0.8 * inch
        use_template = False

        if template_pdf_bytes:
            try:
                top_in, bottom_in = detect_template_margins(template_pdf_bytes)
                top_margin = top_in * inch
                bottom_margin = bottom_in * inch
                use_template = True
            except Exception:
                template_pdf_bytes = None

        pdf_bytes = self._build_content_pdf(
            top_margin=top_margin,
            bottom_margin=bottom_margin,
            use_template=use_template,
        )
        print(f"✅ PDF generado exitosamente ({len(pdf_bytes)} bytes)")

        if template_pdf_bytes:
            try:
                return apply_template_overlay(
                    pdf_bytes,
                    template_pdf_bytes,
                )
            except Exception as exc:
                print(f"⚠️ Error aplicando plantilla institucional: {exc}")

        buffer = BytesIO(pdf_bytes)
        buffer.seek(0)
        return buffer


    
    def generate_excel(self) -> bytes:
        """
        Genera informe en formato Excel (XLSX)
        Usa openpyxl para crear un archivo Excel estructurado
        """
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill
            from openpyxl.utils import get_column_letter
            
            print("📊 Generando informe Excel...")
            
            wb = Workbook()
            
            # HOJA 1: Resumen General
            ws = wb.active
            ws.title = "Resumen General"
            
            # Título
            ws['A1'] = f"INFORME PDM - {self.entity.name}"
            ws['A1'].font = Font(size=16, bold=True)
            anio_texto = "Todos los Años (2024-2027)" if self.anio == 0 else str(self.anio)
            ws['A2'] = f"Año: {anio_texto}"
            ws['A2'].font = Font(size=12)
            
            # Líneas Estratégicas
            ws['A4'] = "AVANCE POR LÍNEAS ESTRATÉGICAS"
            ws['A4'].font = Font(size=14, bold=True)
            
            ws['A5'] = "Línea Estratégica"
            ws['B5'] = "Avance (%)"
            ws['A5'].font = Font(bold=True)
            ws['B5'].font = Font(bold=True)
            
            row = 6
            lineas_data = {}
            for prod in self.productos:
                linea = prod.linea_estrategica or 'Sin Línea'
                if linea not in lineas_data:
                    lineas_data[linea] = {'total': 0, 'suma_avance': 0}
                lineas_data[linea]['total'] += 1
                lineas_data[linea]['suma_avance'] += self.calcular_avance_producto(prod)
            
            for linea, data in lineas_data.items():
                ws[f'A{row}'] = linea
                promedio = data['suma_avance'] / data['total'] if data['total'] > 0 else 0
                ws[f'B{row}'] = f"{promedio:.1f}%"
                row += 1
            
            # HOJA 2: Productos Detallados
            ws2 = wb.create_sheet("Productos")
            ws2['A1'] = "PRODUCTOS Y AVANCES"
            ws2['A1'].font = Font(size=14, bold=True)
            
            headers = ['Código', 'Producto', 'Indicador', 'Meta', 'Unidad', 'Avance Físico', 'Avance Financiero', 'Responsable']
            for col, header in enumerate(headers, 1):
                cell = ws2.cell(row=3, column=col)
                cell.value = header
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="2D3748", end_color="2D3748", fill_type="solid")
                cell.font = Font(bold=True, color="FFFFFF")
            
            row = 4
            for prod in self.productos:
                ws2[f'A{row}'] = prod.codigo_producto
                ws2[f'B{row}'] = prod.producto_mga or 'N/A'
                ws2[f'C{row}'] = prod.indicador_producto_mga or 'N/A'
                ws2[f'D{row}'] = prod.meta_cuatrienio or 0
                ws2[f'E{row}'] = prod.unidad_medida or ''
                ws2[f'F{row}'] = f"{self.calcular_avance_producto(prod):.1f}%"
                ws2[f'G{row}'] = f"{self.calcular_avance_financiero(prod):.1f}%"
                ws2[f'H{row}'] = prod.responsable_secretaria.nombre if prod.responsable_secretaria else 'N/A'
                row += 1
            
            # Ajustar anchos de columna
            for col in range(1, 9):
                ws2.column_dimensions[get_column_letter(col)].width = 20
            
            # HOJA 3: Actividades (mejora implementada)
            ws3 = wb.create_sheet("Actividades")
            ws3['A1'] = "ACTIVIDADES Y ESTADOS"
            ws3['A1'].font = Font(size=14, bold=True)
            
            headers_act = ['Código Producto', 'Actividad', 'Estado', 'Año', 'Meta Ejecutar', 'Fecha Inicio', 'Fecha Fin', 'Responsable', 'Evidencia']
            for col, header in enumerate(headers_act, 1):
                cell = ws3.cell(row=3, column=col)
                cell.value = header
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="2D3748", end_color="2D3748", fill_type="solid")
                cell.font = Font(bold=True, color="FFFFFF")
            
            row = 4
            for act in self.actividades:
                if self.anio == 0 or act.anio == self.anio:
                    ws3[f'A{row}'] = act.clave_producto
                    ws3[f'B{row}'] = act.nombre[:100]
                    ws3[f'C{row}'] = act.estado
                    ws3[f'D{row}'] = act.anio
                    ws3[f'E{row}'] = act.meta_ejecutar or 0
                    ws3[f'F{row}'] = act.fecha_inicio.strftime('%Y-%m-%d') if act.fecha_inicio else ''
                    ws3[f'G{row}'] = act.fecha_fin.strftime('%Y-%m-%d') if act.fecha_fin else ''
                    ws3[f'H{row}'] = act.responsable_secretaria.nombre if act.responsable_secretaria else 'N/A'
                    ws3[f'I{row}'] = 'Sí' if (hasattr(act, 'tiene_evidencia') and act.tiene_evidencia) else 'No'
                    row += 1
            
            # Ajustar anchos
            for col in range(1, 10):
                ws3.column_dimensions[get_column_letter(col)].width = 18
            
            # HOJA 4: Estadísticas de Evidencias (mejora implementada)
            ws4 = wb.create_sheet("Evidencias")
            ws4['A1'] = "RESUMEN DE EVIDENCIAS"
            ws4['A1'].font = Font(size=14, bold=True)
            
            ws4['A3'] = 'Métrica'
            ws4['B3'] = 'Valor'
            ws4['A3'].font = Font(bold=True)
            ws4['B3'].font = Font(bold=True)
            
            total_actividades = len([a for a in self.actividades if self.anio == 0 or a.anio == self.anio])
            actividades_con_evidencia = len([a for a in self.actividades if (self.anio == 0 or a.anio == self.anio) and a.evidencia])
            porcentaje_evidencia = (actividades_con_evidencia / total_actividades * 100) if total_actividades > 0 else 0
            
            ws4['A4'] = 'Total Actividades'
            ws4['B4'] = total_actividades
            ws4['A5'] = 'Actividades con Evidencia'
            ws4['B5'] = actividades_con_evidencia
            ws4['A6'] = 'Porcentaje Documentado'
            ws4['B6'] = f"{porcentaje_evidencia:.1f}%"
            
            ws4['A8'] = 'Productos'
            ws4['B8'] = len(self.productos)
            ws4['A9'] = 'Avance Físico Promedio'
            suma_avances = sum(self.calcular_avance_producto(p) for p in self.productos)
            ws4['B9'] = f"{suma_avances / len(self.productos):.1f}%" if self.productos else "0%"
            
            # Guardar en BytesIO
            from io import BytesIO
            excel_buffer = BytesIO()
            wb.save(excel_buffer)
            excel_bytes = excel_buffer.getvalue()
            excel_buffer.close()
            
            print(f"✅ Excel generado exitosamente con 4 hojas ({len(excel_bytes)} bytes)")
            return excel_bytes
            
        except ImportError as ie:
            print(f"❌ ERROR: Biblioteca no instalada - {ie}")
            raise Exception("El formato Excel no está disponible. Instale openpyxl")
        except Exception as e:
            print(f"❌ Error generando Excel: {e}")
            import traceback
            traceback.print_exc()
            raise
PdmReportGenerator = PDMReportGenerator
