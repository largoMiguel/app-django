"""Generación del Excel Plan de Acción PDM (descarga inmediata, sin persistencia)."""
from __future__ import annotations

from collections import defaultdict
from datetime import datetime
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from apps.entities.models import Entity

from .access import actividades_queryset_for_user, productos_queryset_for_user
from .ejecucion_resumen import ejecucion_por_codigo
from .metrics import (
    ANIOS_PDM,
    actividad_aggs_for_productos,
    estado_producto_anio,
    resumen_anio,
)
from .models import ActividadEstado, PdmProducto

HEADER_FILL = PatternFill(start_color="0E7490", end_color="0E7490", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)

_META_FIELD_BY_ANIO = {
    2024: "programacion_2024",
    2025: "programacion_2025",
    2026: "programacion_2026",
    2027: "programacion_2027",
}

ACTividades_COLUMNS = [
    "DEPENDENCIA",
    "LÍNEA ESTRATÉGICA",
    "CÓDIGO PRODUCTO",
    "PRODUCTO MGA",
    "INDICADOR MGA",
    "UNIDAD MEDIDA",
    "BPIN",
    "META PROGRAMADA (PRODUCTO)",
    "ACTIVIDAD",
    "DESCRIPCIÓN",
    "META ASIGNADA",
    "META EJECUTADA",
    "POR EJECUTAR",
    "ESTADO ACTIVIDAD",
    "RESPONSABLE SECRETARÍA",
    "RESPONSABLE USUARIO",
    "FECHA INICIO",
    "FECHA FIN",
    "EVIDENCIA URL",
    "DESCRIPCIÓN EVIDENCIA",
]

PRODUCTO_COLUMNS = [
    "DEPENDENCIA",
    "LÍNEA ESTRATÉGICA",
    "CÓDIGO PRODUCTO",
    "PRODUCTO MGA",
    "INDICADOR MGA",
    "UNIDAD MEDIDA",
    "BPIN",
    "META PROGRAMADA",
    "META ASIGNADA",
    "META EJECUTADA",
    "POR EJECUTAR",
    "TOTAL ACTIVIDADES",
    "ACTIVIDADES COMPLETADAS",
    "% AVANCE FÍSICO",
    "PRESUPUESTO",
    "PTO. DEFINITIVO",
    "PAGOS",
    "% AVANCE FINANCIERO",
    "ESTADO",
]

DEPENDENCIA_COLUMNS = [
    "DEPENDENCIA",
    "TOTAL PRODUCTOS",
    "META PROGRAMADA",
    "META ASIGNADA",
    "META EJECUTADA",
    "POR EJECUTAR",
    "TOTAL ACTIVIDADES",
    "ACTIVIDADES COMPLETADAS",
    "% AVANCE FÍSICO",
    "PRESUPUESTO",
    "PAGOS",
    "% AVANCE FINANCIERO",
]

ESTADO_LABELS = dict(ActividadEstado.choices)


def _filter_productos_con_meta(qs, anio: int):
    field = _META_FIELD_BY_ANIO.get(anio)
    if not field:
        return qs.none()
    return qs.filter(**{f"{field}__gt": 0})


def _dependencia_nombre(producto: PdmProducto) -> str:
    if producto.responsable_secretaria_id and producto.responsable_secretaria:
        return producto.responsable_secretaria.nombre
    return producto.responsable_secretaria_nombre or "Sin dependencia"


def _pct(numerator: float, denominator: float) -> float:
    if denominator <= 0:
        return 0.0
    return round(min(100.0, (numerator / denominator) * 100), 2)


def _style_header_row(ws, columns: list[str]) -> None:
    for col_idx, title in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=title)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER
    ws.row_dimensions[1].height = 36
    ws.freeze_panes = "A2"


def _write_row(ws, row_idx: int, values: list) -> None:
    for col_idx, val in enumerate(values, start=1):
        cell = ws.cell(row=row_idx, column=col_idx, value=val)
        cell.border = THIN_BORDER
        if isinstance(val, float) and col_idx > 1:
            cell.number_format = "#,##0.00"


def _set_column_widths(ws, widths: list[int]) -> None:
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width


def _gather_data(
    user,
    entity: Entity,
    *,
    anio: int,
    responsable_secretaria_id: int | None = None,
):
    productos_qs = productos_queryset_for_user(user, entity)
    if responsable_secretaria_id:
        productos_qs = productos_qs.filter(responsable_secretaria_id=responsable_secretaria_id)
    productos_qs = _filter_productos_con_meta(productos_qs, anio).select_related(
        "responsable_secretaria", "responsable_usuario"
    )
    productos = list(productos_qs.order_by("linea_estrategica", "codigo_producto", "clave_producto"))
    claves = [p.clave_producto for p in productos]
    codigos = list({p.codigo_producto for p in productos if p.codigo_producto})
    producto_by_clave = {p.clave_producto: p for p in productos}

    aggs = actividad_aggs_for_productos(entity.id, claves)
    ejecucion = ejecucion_por_codigo(entity.id, codigos, anio)

    act_qs = actividades_queryset_for_user(user, entity).filter(anio=anio)
    if responsable_secretaria_id:
        act_qs = act_qs.filter(clave_producto__in=claves)
    actividades = list(
        act_qs.select_related("responsable_secretaria", "responsable_usuario")
        .prefetch_related("evidencia")
        .order_by("clave_producto", "id")
    )

    return {
        "productos": productos,
        "producto_by_clave": producto_by_clave,
        "aggs": aggs,
        "ejecucion": ejecucion,
        "actividades": actividades,
    }


def _build_actividades_sheet(ws, data: dict, anio: int) -> None:
    _style_header_row(ws, ACTividades_COLUMNS)
    row_idx = 2
    producto_by_clave = data["producto_by_clave"]

    for act in data["actividades"]:
        producto = producto_by_clave.get(act.clave_producto)
        if producto is None:
            continue

        resumen = resumen_anio(producto, anio, data["aggs"].get(act.clave_producto, {}))
        meta_asignada = float(act.meta_ejecutar or 0)
        meta_ejecutada = meta_asignada if act.estado == ActividadEstado.COMPLETADA else 0.0
        por_ejecutar = max(0.0, meta_asignada - meta_ejecutada)

        evidencia = getattr(act, "evidencia", None)
        sec_nombre = ""
        if act.responsable_secretaria_id and act.responsable_secretaria:
            sec_nombre = act.responsable_secretaria.nombre
        elif producto.responsable_secretaria_id and producto.responsable_secretaria:
            sec_nombre = producto.responsable_secretaria.nombre
        else:
            sec_nombre = producto.responsable_secretaria_nombre or ""

        usuario_nombre = ""
        if act.responsable_usuario_id and act.responsable_usuario:
            usuario_nombre = act.responsable_usuario.full_name or act.responsable_usuario.email

        values = [
            _dependencia_nombre(producto),
            producto.linea_estrategica or "",
            producto.codigo_producto or "",
            producto.producto_mga or "",
            producto.indicador_producto_mga or "",
            producto.unidad_medida or "",
            producto.bpin or "",
            resumen["meta_programada"],
            act.nombre,
            act.descripcion or "",
            meta_asignada,
            meta_ejecutada,
            por_ejecutar,
            ESTADO_LABELS.get(act.estado, act.estado),
            sec_nombre,
            usuario_nombre,
            act.fecha_inicio.date().isoformat() if act.fecha_inicio else "",
            act.fecha_fin.date().isoformat() if act.fecha_fin else "",
            evidencia.url_evidencia if evidencia else "",
            evidencia.descripcion if evidencia else "",
        ]
        _write_row(ws, row_idx, values)
        row_idx += 1

    _set_column_widths(
        ws,
        [28, 30, 14, 40, 35, 14, 16, 16, 40, 35, 14, 14, 14, 16, 28, 28, 12, 12, 40, 40],
    )


def _build_producto_sheet(ws, data: dict, anio: int) -> None:
    _style_header_row(ws, PRODUCTO_COLUMNS)
    row_idx = 2

    for producto in data["productos"]:
        aggs_by_anio = data["aggs"].get(producto.clave_producto, {})
        resumen = resumen_anio(producto, anio, aggs_by_anio)
        ej = data["ejecucion"].get(str(producto.codigo_producto), {})
        pto_def = float(ej.get("pto_definitivo", 0))
        pagos = float(ej.get("pagos", 0))
        presupuesto = resumen["presupuesto"]
        por_ejecutar = max(0.0, resumen["meta_programada"] - resumen["meta_ejecutada"])

        values = [
            _dependencia_nombre(producto),
            producto.linea_estrategica or "",
            producto.codigo_producto or "",
            producto.producto_mga or "",
            producto.indicador_producto_mga or "",
            producto.unidad_medida or "",
            producto.bpin or "",
            resumen["meta_programada"],
            resumen["meta_asignada"],
            resumen["meta_ejecutada"],
            por_ejecutar,
            resumen["total_actividades"],
            resumen["actividades_completadas"],
            resumen["porcentaje_avance"],
            presupuesto,
            pto_def,
            pagos,
            _pct(pagos, pto_def if pto_def > 0 else presupuesto),
            estado_producto_anio(producto, anio, aggs_by_anio),
        ]
        _write_row(ws, row_idx, values)
        row_idx += 1

    _set_column_widths(
        ws,
        [28, 30, 14, 40, 35, 14, 16, 14, 14, 14, 14, 14, 14, 14, 16, 16, 16, 14, 14],
    )


def _build_dependencia_sheet(ws, data: dict, anio: int) -> None:
    _style_header_row(ws, DEPENDENCIA_COLUMNS)
    grouped: dict[str, dict] = defaultdict(
        lambda: {
            "productos": 0,
            "meta_programada": 0.0,
            "meta_asignada": 0.0,
            "meta_ejecutada": 0.0,
            "total_actividades": 0,
            "actividades_completadas": 0,
            "presupuesto": 0.0,
            "pagos": 0.0,
            "avance_sum": 0.0,
            "avance_count": 0,
        }
    )

    for producto in data["productos"]:
        dep = _dependencia_nombre(producto)
        aggs_by_anio = data["aggs"].get(producto.clave_producto, {})
        resumen = resumen_anio(producto, anio, aggs_by_anio)
        ej = data["ejecucion"].get(str(producto.codigo_producto), {})
        g = grouped[dep]
        g["productos"] += 1
        g["meta_programada"] += resumen["meta_programada"]
        g["meta_asignada"] += resumen["meta_asignada"]
        g["meta_ejecutada"] += resumen["meta_ejecutada"]
        g["total_actividades"] += resumen["total_actividades"]
        g["actividades_completadas"] += resumen["actividades_completadas"]
        g["presupuesto"] += resumen["presupuesto"]
        g["pagos"] += float(ej.get("pagos", 0))
        if resumen["meta_programada"] > 0:
            g["avance_sum"] += resumen["porcentaje_avance"]
            g["avance_count"] += 1

    row_idx = 2
    for dep in sorted(grouped.keys()):
        g = grouped[dep]
        por_ejecutar = max(0.0, g["meta_programada"] - g["meta_ejecutada"])
        avance_fisico = round(g["avance_sum"] / g["avance_count"], 2) if g["avance_count"] else 0.0
        presupuesto = g["presupuesto"]
        values = [
            dep,
            g["productos"],
            g["meta_programada"],
            g["meta_asignada"],
            g["meta_ejecutada"],
            por_ejecutar,
            g["total_actividades"],
            g["actividades_completadas"],
            avance_fisico,
            presupuesto,
            g["pagos"],
            _pct(g["pagos"], presupuesto),
        ]
        _write_row(ws, row_idx, values)
        row_idx += 1

    _set_column_widths(ws, [32, 14, 16, 16, 16, 16, 16, 16, 14, 16, 16, 14])


def build_plan_accion_workbook(
    entity: Entity,
    user,
    anio: int,
    *,
    responsable_secretaria_id: int | None = None,
) -> Workbook:
    if anio not in ANIOS_PDM:
        raise ValueError(f"Año no válido: {anio}")

    data = _gather_data(
        user,
        entity,
        anio=anio,
        responsable_secretaria_id=responsable_secretaria_id,
    )

    wb = Workbook()
    ws_act = wb.active
    ws_act.title = "Plan de acción"
    _build_actividades_sheet(ws_act, data, anio)

    ws_prod = wb.create_sheet("Resumen por producto")
    _build_producto_sheet(ws_prod, data, anio)

    ws_dep = wb.create_sheet("Resumen por dependencia")
    _build_dependencia_sheet(ws_dep, data, anio)

    return wb


def build_plan_accion_export(
    entity: Entity,
    user,
    anio: int,
    *,
    responsable_secretaria_id: int | None = None,
) -> tuple[bytes, str]:
    wb = build_plan_accion_workbook(
        entity,
        user,
        anio,
        responsable_secretaria_id=responsable_secretaria_id,
    )
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    dep_suffix = f"_dep{responsable_secretaria_id}" if responsable_secretaria_id else ""
    filename = f"Plan_Accion_PDM_{entity.slug or entity.id}_{anio}{dep_suffix}.xlsx"
    return buffer.getvalue(), filename
