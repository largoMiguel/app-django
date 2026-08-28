"""Generación del Excel de exportación PIIP (sin persistencia)."""
from __future__ import annotations

import re
from collections import defaultdict
from io import BytesIO

from django.contrib.auth import get_user_model
from django.db.models import Sum
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from apps.entities.models import Entity

from .access import productos_queryset_for_user
from .bpin_view import consultar_bpines_externos
from .ejecucion_mensual import ejecucion_mensual_por_producto_fuente
from .fuente_financiacion import normalizar_fuente_piip
from .metrics import ANIOS_PDM, actividad_aggs_for_productos
from .models import PDMEjecucionPresupuestal, PdmProducto

User = get_user_model()

_META_FIELD_BY_ANIO = {
    2024: "programacion_2024",
    2025: "programacion_2025",
    2026: "programacion_2026",
    2027: "programacion_2027",
}

PIIP_COLUMNS = [
    "CODIGO BPIN",
    "PROYECTO PIIP",
    "SECTOR",
    "CÓDIGO DE PRODUCTO",
    "META (PRODUCTO) ASIGNADO",
    "CANTIDAD META PROGRAMADA",
    "CANTIDAD META EJECUTADA",
    "VALOR INICIAL",
    " Comprometido cada mes ",
    " Pago Cada mes ",
    " Comprometido Acumulado ",
    " Pago Acumulado ",
    "VALOR EJECUTADO",
    "FUENTE DE FINANCIACIÓN",
    "RESPONSABLE",
]

HEADER_FILL = PatternFill(start_color="6AA84F", end_color="6AA84F", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)

COLUMN_WIDTHS = {
    "A": 18,
    "B": 55,
    "C": 28,
    "D": 16,
    "E": 45,
    "F": 14,
    "G": 14,
    "H": 18,
    "I": 18,
    "J": 14,
    "K": 20,
    "L": 16,
    "M": 18,
    "N": 35,
    "O": 45,
}


def _split_bpines(bpin_raw: str | None) -> list[str]:
    """Separa BPINs múltiples (coma o punto y coma) en códigos individuales."""
    if not bpin_raw:
        return []
    out: list[str] = []
    seen: set[str] = set()
    for part in re.split(r"[,;]+", str(bpin_raw)):
        bpin = part.strip()
        if bpin and bpin not in seen:
            seen.add(bpin)
            out.append(bpin)
    return out


def _meta_programada(producto: PdmProducto, anio: int) -> float:
    field = _META_FIELD_BY_ANIO.get(anio)
    if not field:
        return 0.0
    return float(getattr(producto, field, 0) or 0)


def _filter_productos_piip(qs, anio: int):
    field = _META_FIELD_BY_ANIO.get(anio)
    if not field:
        return qs.none()
    return (
        qs.exclude(bpin__isnull=True)
        .exclude(bpin="")
        .filter(**{f"{field}__gt": 0})
        .order_by("bpin", "codigo_producto")
    )


def _fuentes_detalle_por_producto(
    entity_id: int,
    codigos: list[str],
    anio: int,
) -> dict[str, list[dict[str, float | str]]]:
    """Ejecución por producto y fuente (normalizada), agregando duplicados del mismo nombre."""
    if not codigos:
        return {}

    rows = (
        PDMEjecucionPresupuestal.objects.filter(
            entity_id=entity_id,
            codigo_producto__in=codigos,
            anio=anio,
        )
        .values("codigo_producto", "descripcion_fte")
        .annotate(
            pto_definitivo=Sum("pto_definitivo"),
            pagos=Sum("pagos"),
        )
    )

    grouped: dict[str, dict[str, dict[str, float | str]]] = defaultdict(dict)
    for row in rows:
        codigo = row["codigo_producto"]
        nombre = normalizar_fuente_piip(row["descripcion_fte"])
        bucket = grouped[codigo].setdefault(
            nombre,
            {"nombre": nombre, "pto_definitivo": 0.0, "pagos": 0.0},
        )
        bucket["pto_definitivo"] = float(bucket["pto_definitivo"]) + float(row["pto_definitivo"] or 0)
        bucket["pagos"] = float(bucket["pagos"]) + float(row["pagos"] or 0)

    return {
        codigo: sorted(fuentes.values(), key=lambda x: str(x["nombre"]))
        for codigo, fuentes in grouped.items()
    }


def _secretarios_por_secretaria(entity_id: int, secretaria_ids: list[int]) -> dict[int, str]:
    if not secretaria_ids:
        return {}
    users = User.objects.filter(
        entity_id=entity_id,
        role="secretario",
        secretaria_id__in=secretaria_ids,
        is_active=True,
    ).values("secretaria_id", "full_name")
    out: dict[int, str] = {}
    for row in users:
        sid = row["secretaria_id"]
        if sid and sid not in out:
            name = (row["full_name"] or "").strip()
            if name:
                out[sid] = name
    return out


def _format_responsable(
    secretario_nombre: str | None,
    secretaria_nombre: str | None,
) -> str:
    sec = (secretaria_nombre or "").strip()
    secr = (secretario_nombre or "").strip()
    if secr and sec:
        return f"{secr} - {sec}".upper()
    if secr:
        return secr.upper()
    if sec:
        return sec.upper()
    return ""


def build_piip_export_rows(entity: Entity, user, anio: int, mes: int | None = None) -> list[list]:
    """Arma filas: una por cada BPIN (separados por coma) y por cada fuente presupuestal."""
    if anio not in ANIOS_PDM:
        return []

    qs = _filter_productos_piip(productos_queryset_for_user(user, entity), anio)
    productos = list(qs)
    if not productos:
        return []

    codigos = [p.codigo_producto for p in productos]
    claves = [p.clave_producto for p in productos]
    all_bpines: list[str] = []
    seen_bpin: set[str] = set()
    for p in productos:
        for bpin in _split_bpines(p.bpin):
            if bpin not in seen_bpin:
                seen_bpin.add(bpin)
                all_bpines.append(bpin)

    aggs_map = actividad_aggs_for_productos(
        entity.id,
        claves,
        hasta_mes=mes,
        anio_corte=anio if mes else None,
    )
    fuentes_map = _fuentes_detalle_por_producto(entity.id, codigos, anio)
    fuentes_mensual_map = (
        ejecucion_mensual_por_producto_fuente(entity.id, codigos, anio, mes) if mes else {}
    )
    datos_abiertos, _ = consultar_bpines_externos(all_bpines)

    secretaria_ids = [p.responsable_secretaria_id for p in productos if p.responsable_secretaria_id]
    secretarios_map = _secretarios_por_secretaria(entity.id, secretaria_ids)

    rows: list[list] = []
    for producto in productos:
        bpines = _split_bpines(producto.bpin)
        if not bpines:
            continue

        aggs_anio = aggs_map.get(producto.clave_producto, {}).get(anio, {})
        meta_ejecutada = float(aggs_anio.get("meta_ejecutada", 0) or 0)
        meta_programada = _meta_programada(producto, anio)

        secretaria_nombre = None
        if producto.responsable_secretaria:
            secretaria_nombre = producto.responsable_secretaria.nombre
        elif producto.responsable_secretaria_nombre:
            secretaria_nombre = producto.responsable_secretaria_nombre

        secretario_nombre = None
        if producto.responsable_secretaria_id:
            secretario_nombre = secretarios_map.get(producto.responsable_secretaria_id)

        responsable = _format_responsable(secretario_nombre, secretaria_nombre)

        fuentes_list = fuentes_map.get(producto.codigo_producto) or [
            {"nombre": "Otros", "pto_definitivo": 0.0, "pagos": 0.0},
        ]
        fuentes_mes_list = fuentes_mensual_map.get(producto.codigo_producto) or []
        fuentes_mes_by_name = {str(f["nombre"]): f for f in fuentes_mes_list}

        for bpin in bpines:
            externo = datos_abiertos.get(bpin) or {}
            nombre_proyecto = externo.get("nombreproyecto") or ""
            sector = externo.get("sector") or ""

            for fuente in fuentes_list:
                nombre_fuente = str(fuente["nombre"])
                fm = fuentes_mes_by_name.get(nombre_fuente, {})
                comprometido_mes = float(fm.get("saldo_mes", 0) or 0) if mes else 0
                pago_mes = float(fm.get("pagos_mes", 0) or 0) if mes else 0
                comprometido_acum = float(fm.get("saldo_acum", 0) or 0) if mes else 0
                pago_acum = float(fm.get("pagos_acum", 0) or 0) if mes else 0
                valor_ejecutado = pago_acum if mes else float(fuente["pagos"])

                rows.append(
                    [
                        bpin,
                        nombre_proyecto,
                        sector,
                        producto.codigo_producto,
                        producto.producto_mga or "",
                        meta_programada,
                        meta_ejecutada,
                        float(fuente["pto_definitivo"]),
                        comprometido_mes,
                        pago_mes,
                        comprometido_acum,
                        pago_acum,
                        valor_ejecutado,
                        nombre_fuente,
                        responsable,
                    ]
                )
    return rows


def build_piip_workbook(entity: Entity, user, anio: int, mes: int | None = None) -> Workbook:
    """Construye el workbook Excel con estilo de encabezado verde."""
    wb = Workbook()
    ws = wb.active
    title = f"PIIP {anio}"
    if mes:
        title = f"PIIP {anio}-{mes:02d}"
    ws.title = title

    for col_idx, header in enumerate(PIIP_COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER

    for row_idx, row_data in enumerate(build_piip_export_rows(entity, user, anio, mes=mes), start=2):
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = THIN_BORDER
            if col_idx in (6, 7):
                cell.alignment = Alignment(horizontal="center")
            elif col_idx in (8, 9, 10, 11, 12, 13):
                cell.number_format = "#,##0.00"

    ws.row_dimensions[1].height = 30
    for col_letter, width in COLUMN_WIDTHS.items():
        ws.column_dimensions[col_letter].width = width

    ws.freeze_panes = "A2"
    return wb


def workbook_to_bytes(wb: Workbook) -> bytes:
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()
