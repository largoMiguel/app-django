"""Exportación Excel — informe trimestral Planes Institucionales."""
from __future__ import annotations

from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from apps.entities.models import Entity

from .access import actividades_queryset_for_user, planes_queryset_for_user
from .models import Trimestre

HEADER_FILL = PatternFill(start_color="3eafd4", end_color="3eafd4", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)

COLUMNS = [
    "PLAN",
    "CÓDIGO",
    "VIGENCIA",
    "TRIMESTRE",
    "ACTIVIDAD",
    "META",
    "INDICADOR",
    "RESPONSABLE SECRETARÍA",
    "ESTADO",
    "AVANCE %",
    "FECHA INICIO",
    "FECHA FIN",
    "EVIDENCIA URL",
    "DESCRIPCIÓN EVIDENCIA",
]


def build_trimestral_excel(
    user,
    entity: Entity,
    *,
    anio: int,
    trimestre: int | None = None,
    plan_id: int | None = None,
    responsable_secretaria_id: int | None = None,
) -> tuple[BytesIO, str]:
    act_qs = (
        actividades_queryset_for_user(user, entity)
        .filter(anio=anio)
        .select_related("plan", "plan__catalogo", "responsable_secretaria", "evidencia")
        .order_by("plan__catalogo__orden", "plan_id", "trimestre", "id")
    )
    if trimestre:
        act_qs = act_qs.filter(trimestre=trimestre)
    if plan_id:
        act_qs = act_qs.filter(plan_id=plan_id)
    if responsable_secretaria_id:
        act_qs = act_qs.filter(responsable_secretaria_id=responsable_secretaria_id)

    wb = Workbook()
    ws = wb.active
    ws.title = "Informe trimestral"

    for col_idx, title in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=title)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER

    row_idx = 2
    for act in act_qs:
        evidencia = getattr(act, "evidencia", None)
        tri_label = Trimestre(act.trimestre).label if act.trimestre in Trimestre.values else str(act.trimestre)
        values = [
            act.plan.nombre,
            act.plan.catalogo.codigo,
            act.anio,
            tri_label,
            act.nombre,
            act.meta,
            act.indicador,
            act.responsable_secretaria.nombre if act.responsable_secretaria_id else "",
            act.get_estado_display(),
            act.avance,
            act.fecha_inicio.isoformat() if act.fecha_inicio else "",
            act.fecha_fin.isoformat() if act.fecha_fin else "",
            evidencia.url_evidencia if evidencia else "",
            evidencia.descripcion if evidencia else "",
        ]
        for col_idx, val in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border = THIN_BORDER
        row_idx += 1

    widths = [40, 14, 10, 14, 45, 30, 30, 28, 14, 10, 12, 12, 40, 50]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + idx) if idx <= 26 else "A"].width = width

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    tri_suffix = f"_T{trimestre}" if trimestre else ""
    filename = f"Planes_D612_{entity.slug or entity.id}_{anio}{tri_suffix}.xlsx"
    return buf, filename
